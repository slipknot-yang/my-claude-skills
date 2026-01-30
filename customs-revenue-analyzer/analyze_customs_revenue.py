#!/usr/bin/env python3
"""
관세 수입 현황 다각도 분석 스크립트

Oracle DB에서 관세 데이터를 추출하여 Excel 파일로 내보냅니다.
Claude for Excel에서 추가 분석 및 시각화가 가능합니다.

Usage:
    python analyze_customs_revenue.py
    python analyze_customs_revenue.py --output /path/to/output.xlsx
    python analyze_customs_revenue.py --year 2024
"""

import oracledb
import pandas as pd
from datetime import datetime
import argparse
import sys

# DB 접속 정보
DB_CONFIG = {
    "user": "CLRIUSR",
    "password": "ntancisclri1!",
    "dsn": "211.239.120.42:3535/NTANCIS"
}

# HS 코드 매핑
HS2_NAMES = {
    '74': '구리와 그 제품',
    '27': '광물성 연료, 광물유',
    '87': '차량(철도/전차 제외)',
    '26': '광, 슬래그 및 회',
    '84': '원자로, 보일러, 기계류',
    '85': '전기기기',
    '39': '플라스틱과 그 제품',
    '72': '철강',
    '24': '담배',
    '73': '철강 제품',
    '63': '방직용 섬유제품',
    '28': '무기화학품',
    '15': '동물성/식물성 유지',
    '40': '고무와 그 제품',
    '17': '당류와 설탕과자',
    '94': '가구, 침구',
    '10': '곡물',
    '52': '면',
    '22': '음료, 알코올',
    '08': '과실과 견과류',
}

# 국가 코드 매핑
COUNTRY_NAMES = {
    'CD': '콩고민주공화국',
    'TZ': '탄자니아',
    'ZM': '잠비아',
    'AE': '아랍에미리트',
    'CN': '중국',
    'IN': '인도',
    'JP': '일본',
    'ZA': '남아프리카공화국',
    'SA': '사우디아라비아',
    'US': '미국',
    'KE': '케냐',
    'KR': '한국',
    'DE': '독일',
    'CH': '스위스',
    'UG': '우간다',
}


def connect_db():
    """Oracle DB 연결"""
    print("🔗 DB 연결 중...")
    try:
        conn = oracledb.connect(**DB_CONFIG)
        print("✅ DB 연결 성공")
        return conn
    except oracledb.Error as e:
        print(f"❌ DB 연결 실패: {e}")
        sys.exit(1)


def fetch_yearly_data(conn):
    """연도별 관세 수입 추출"""
    print("📊 연도별 데이터 추출 중...")
    query = """
    SELECT 
        '20' || TANSAD_YY as YEAR,
        COUNT(*) as ITEM_COUNT,
        SUM(ITM_TAX_AMT) as TOTAL_TAX,
        SUM(ITM_INVC_USD_AMT) as TOTAL_VALUE_USD,
        ROUND(AVG(ITM_TAX_AMT), 0) as AVG_TAX
    FROM CLRI_TANSAD_ITM_D
    WHERE DEL_YN = 'N' AND TANSAD_YY >= '20'
    GROUP BY TANSAD_YY
    ORDER BY TANSAD_YY DESC
    """
    df = pd.read_sql(query, conn)
    
    # 성장률 계산
    df['GROWTH_RATE'] = df['TOTAL_TAX'].pct_change(-1) * 100
    df['GROWTH_RATE'] = df['GROWTH_RATE'].round(1)
    
    print(f"  → {len(df)}개 연도 데이터")
    return df


def fetch_commodity_data(conn):
    """품목별(HS2) 관세 수입 추출"""
    print("📦 품목별 데이터 추출 중...")
    query = """
    SELECT 
        SUBSTR(HS_CD, 1, 2) as HS2_CODE,
        COUNT(*) as ITEM_COUNT,
        SUM(ITM_TAX_AMT) as TOTAL_TAX,
        SUM(ITM_INVC_USD_AMT) as TOTAL_VALUE_USD,
        ROUND(AVG(ITM_TAX_AMT), 0) as AVG_TAX
    FROM CLRI_TANSAD_ITM_D
    WHERE DEL_YN = 'N' AND ITM_TAX_AMT > 0
    GROUP BY SUBSTR(HS_CD, 1, 2)
    ORDER BY TOTAL_TAX DESC
    FETCH FIRST 30 ROWS ONLY
    """
    df = pd.read_sql(query, conn)
    
    # 품목명 추가
    df['HS2_NAME'] = df['HS2_CODE'].map(HS2_NAMES).fillna('기타')
    
    # 비중 계산
    total = df['TOTAL_TAX'].sum()
    df['TAX_SHARE'] = (df['TOTAL_TAX'] / total * 100).round(1)
    
    print(f"  → {len(df)}개 품목 데이터")
    return df


def fetch_country_data(conn):
    """국가별 수입 현황 추출"""
    print("🌍 국가별 데이터 추출 중...")
    query = """
    SELECT 
        ORIG_CNTY_CD as COUNTRY_CODE,
        COUNT(*) as ITEM_COUNT,
        SUM(ITM_TAX_AMT) as TOTAL_TAX,
        SUM(ITM_INVC_USD_AMT) as TOTAL_VALUE_USD
    FROM CLRI_TANSAD_ITM_D
    WHERE DEL_YN = 'N' AND ORIG_CNTY_CD IS NOT NULL
    GROUP BY ORIG_CNTY_CD
    ORDER BY TOTAL_VALUE_USD DESC NULLS LAST
    FETCH FIRST 30 ROWS ONLY
    """
    df = pd.read_sql(query, conn)
    
    # 국가명 추가
    df['COUNTRY_NAME'] = df['COUNTRY_CODE'].map(COUNTRY_NAMES).fillna('기타')
    
    # 비중 계산
    total = df['TOTAL_VALUE_USD'].sum()
    df['VALUE_SHARE'] = (df['TOTAL_VALUE_USD'] / total * 100).round(1)
    
    print(f"  → {len(df)}개 국가 데이터")
    return df


def fetch_customs_office_data(conn):
    """세관별 현황 추출"""
    print("🏛️ 세관별 데이터 추출 중...")
    query = """
    SELECT 
        CSTM_OFCE_CD as CUSTOMS_OFFICE,
        COUNT(*) as ITEM_COUNT,
        SUM(ITM_TAX_AMT) as TOTAL_TAX,
        SUM(ITM_INVC_USD_AMT) as TOTAL_VALUE_USD
    FROM CLRI_TANSAD_ITM_D
    WHERE DEL_YN = 'N'
    GROUP BY CSTM_OFCE_CD
    ORDER BY TOTAL_TAX DESC NULLS LAST
    """
    df = pd.read_sql(query, conn)
    print(f"  → {len(df)}개 세관 데이터")
    return df


def fetch_monthly_data(conn):
    """월별 추이 추출 (최근 24개월)"""
    print("📅 월별 데이터 추출 중...")
    query = """
    SELECT 
        TO_CHAR(FRST_RGSR_DTM, 'YYYY-MM') as MONTH,
        COUNT(*) as ITEM_COUNT,
        SUM(ITM_TAX_AMT) as TOTAL_TAX,
        SUM(ITM_INVC_USD_AMT) as TOTAL_VALUE_USD
    FROM CLRI_TANSAD_ITM_D
    WHERE DEL_YN = 'N' 
      AND FRST_RGSR_DTM >= ADD_MONTHS(SYSDATE, -24)
    GROUP BY TO_CHAR(FRST_RGSR_DTM, 'YYYY-MM')
    ORDER BY MONTH
    """
    df = pd.read_sql(query, conn)
    print(f"  → {len(df)}개 월 데이터")
    return df


def create_summary(df_yearly, df_commodity, df_country):
    """요약 데이터 생성"""
    summary = {
        '지표': [
            '분석 기간',
            '총 건수',
            '총 세액 (현지화)',
            '총 수입액 (USD)',
            'TOP 품목',
            'TOP 교역국',
            '평균 연간 세액',
            '최근 성장률'
        ],
        '값': [
            f"{df_yearly['YEAR'].min()} ~ {df_yearly['YEAR'].max()}",
            f"{df_yearly['ITEM_COUNT'].sum():,.0f} 건",
            f"{df_yearly['TOTAL_TAX'].sum():,.0f}",
            f"${df_yearly['TOTAL_VALUE_USD'].sum():,.0f}",
            f"{df_commodity.iloc[0]['HS2_CODE']} ({df_commodity.iloc[0]['HS2_NAME']})",
            f"{df_country.iloc[0]['COUNTRY_CODE']} ({df_country.iloc[0]['COUNTRY_NAME']})",
            f"{df_yearly['TOTAL_TAX'].mean():,.0f}",
            f"{df_yearly['GROWTH_RATE'].iloc[0]:+.1f}%" if pd.notna(df_yearly['GROWTH_RATE'].iloc[0]) else "N/A"
        ]
    }
    return pd.DataFrame(summary)


def create_claude_prompts():
    """Claude for Excel 분석 프롬프트 생성"""
    prompts = {
        '시트명': [
            '요약',
            '연도별_추이',
            '품목별_현황',
            '국가별_현황',
            '월별_추이'
        ],
        'Claude 프롬프트': [
            '이 관세 수입 데이터의 주요 인사이트를 3가지로 요약해주세요.',
            '연도별 관세 수입 추이를 분석하고, 성장률 변화의 원인과 향후 전망을 예측해주세요.',
            'TOP 10 품목의 관세 수입 특성을 분석하고, 각 품목별 수입 트렌드를 설명해주세요.',
            '주요 교역국의 수입 패턴을 분석하고, 국가별 특징과 리스크를 평가해주세요.',
            '월별 변동 패턴에서 계절성이 있는지 분석하고, 이상치가 있다면 원인을 추정해주세요.'
        ],
        'Excel 함수': [
            '=CLAUDE("요약 분석", A1:B10)',
            '=CLAUDE("연도별 추이 분석", A1:G10)',
            '=CLAUDE("품목별 분석", A1:G30)',
            '=CLAUDE("국가별 분석", A1:F30)',
            '=CLAUDE("월별 추이 분석", A1:D25)'
        ]
    }
    return pd.DataFrame(prompts)


def save_to_excel(output_path, df_yearly, df_commodity, df_country, df_customs, df_monthly):
    """Excel 파일 저장"""
    print(f"\n📁 Excel 파일 생성 중: {output_path}")
    
    # 요약 데이터 생성
    df_summary = create_summary(df_yearly, df_commodity, df_country)
    df_prompts = create_claude_prompts()
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='요약', index=False)
        df_yearly.to_excel(writer, sheet_name='연도별_추이', index=False)
        df_commodity.to_excel(writer, sheet_name='품목별_현황', index=False)
        df_country.to_excel(writer, sheet_name='국가별_현황', index=False)
        df_customs.to_excel(writer, sheet_name='세관별_현황', index=False)
        df_monthly.to_excel(writer, sheet_name='월별_추이', index=False)
        df_prompts.to_excel(writer, sheet_name='Claude_분석_가이드', index=False)
    
    print(f"✅ Excel 파일 저장 완료")


def add_charts(output_path):
    """Excel에 차트 추가"""
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        
        print("📈 차트 추가 중...")
        wb = load_workbook(output_path)
        
        # 연도별 추이 - 막대 차트
        ws = wb['연도별_추이']
        chart = BarChart()
        chart.title = "연도별 관세 수입 추이"
        chart.x_axis.title = "연도"
        chart.y_axis.title = "세액"
        chart.height = 12
        chart.width = 18
        
        data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row, max_col=3)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "I2")
        
        # 월별 추이 - 라인 차트
        ws = wb['월별_추이']
        if ws.max_row > 1:
            chart = LineChart()
            chart.title = "월별 관세 수입 추이"
            chart.x_axis.title = "월"
            chart.y_axis.title = "세액"
            chart.height = 12
            chart.width = 20
            
            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row, max_col=3)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "F2")
        
        wb.save(output_path)
        print("✅ 차트 추가 완료")
        
    except ImportError:
        print("⚠️ openpyxl 미설치로 차트 생성 생략")


def print_summary(df_yearly, df_commodity, df_country):
    """콘솔에 요약 출력"""
    print("\n" + "="*60)
    print("📊 관세 수입 현황 분석 요약")
    print("="*60)
    
    print(f"\n📅 분석 기간: {df_yearly['YEAR'].min()} ~ {df_yearly['YEAR'].max()}")
    print(f"📦 총 건수: {df_yearly['ITEM_COUNT'].sum():,.0f} 건")
    print(f"💰 총 세액: {df_yearly['TOTAL_TAX'].sum():,.0f}")
    print(f"💵 총 수입액: ${df_yearly['TOTAL_VALUE_USD'].sum():,.0f}")
    
    print("\n🏆 TOP 5 품목 (세액 기준):")
    for i, row in df_commodity.head(5).iterrows():
        print(f"  {i+1}. HS {row['HS2_CODE']} ({row['HS2_NAME']}): {row['TOTAL_TAX']:,.0f} ({row['TAX_SHARE']}%)")
    
    print("\n🌍 TOP 5 교역국 (수입액 기준):")
    for i, row in df_country.head(5).iterrows():
        print(f"  {i+1}. {row['COUNTRY_CODE']} ({row['COUNTRY_NAME']}): ${row['TOTAL_VALUE_USD']:,.0f}")
    
    print("\n" + "="*60)


def main():
    parser = argparse.ArgumentParser(description='관세 수입 현황 분석')
    parser.add_argument('--output', '-o', default='customs_revenue_analysis.xlsx',
                        help='출력 Excel 파일 경로')
    parser.add_argument('--year', '-y', type=str, default=None,
                        help='특정 연도만 분석 (예: 2024)')
    args = parser.parse_args()
    
    print("🚀 관세 수입 현황 분석 시작")
    print(f"⏰ 시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # DB 연결
    conn = connect_db()
    
    try:
        # 데이터 추출
        df_yearly = fetch_yearly_data(conn)
        df_commodity = fetch_commodity_data(conn)
        df_country = fetch_country_data(conn)
        df_customs = fetch_customs_office_data(conn)
        df_monthly = fetch_monthly_data(conn)
        
        # Excel 저장
        save_to_excel(args.output, df_yearly, df_commodity, df_country, df_customs, df_monthly)
        
        # 차트 추가
        add_charts(args.output)
        
        # 요약 출력
        print_summary(df_yearly, df_commodity, df_country)
        
        print(f"\n✅ 분석 완료!")
        print(f"📁 결과 파일: {args.output}")
        print(f"\n💡 Claude for Excel에서 'Claude_분석_가이드' 시트의 프롬프트를 활용하세요.")
        
    finally:
        conn.close()
        print("🔌 DB 연결 종료")


if __name__ == "__main__":
    main()
