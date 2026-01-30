#!/usr/bin/env python3
"""
프리미엄 관세 분석 보고서 생성기 (한국어 버전)

WCO PMM / KCS 관세연감 / UN Comtrade 수준의 보고서 생성

Features:
- 경영진 대시보드 (Executive Summary)
- WCO PMM KPI 스코어카드
- 고급 시각화 (파레토, 히트맵, 리스크 매트릭스)
- 전문 포맷팅
- 분석방법론 & 용어정의

Usage:
    python premium_reports_kr.py
"""

import oracledb
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys

# 로컬 모듈 임포트
from kpi_calculator import KPICalculator, KPI_DEFINITIONS, KPICategory, format_currency, format_percent
from visualizations import (
    ColorPalette, StyleManager, 
    add_kpi_card, add_risk_matrix, add_scorecard_table,
    add_heatmap_formatting, add_databar_formatting,
    add_pareto_chart, add_combo_chart,
    write_styled_dataframe, get_trend_arrow
)

# === 설정 ===
DB_CONFIG = {
    "user": "CLRIUSR",
    "password": "ntancisclri1!",
    "dsn": "211.239.120.42:3535/NTANCIS"
}

BASE_PATH = os.path.dirname(os.path.abspath(__file__))


class PremiumReportGeneratorKR:
    """프리미엄 관세 보고서 생성기 (한국어)"""
    
    def __init__(self, conn):
        self.conn = conn
        self.kpi_calc = KPICalculator(conn)
        self.sm = StyleManager()
        self.report_date = datetime.now().strftime('%Y년 %m월 %d일')
    
    # === 공통 헬퍼 함수 ===
    
    def _create_cover_sheet(self, wb: Workbook, title: str, subtitle: str, metrics: dict) -> Worksheet:
        """표지 시트 생성"""
        ws = wb.create_sheet("표지", 0)
        
        # 배경색
        for row in range(1, 35):
            for col in range(1, 16):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'
                )
        
        # 상단 배너
        for col in range(1, 16):
            for row in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=ColorPalette.PRIMARY, 
                    end_color=ColorPalette.PRIMARY, 
                    fill_type='solid'
                )
        
        # 로고 텍스트
        ws.merge_cells('B2:N2')
        logo_cell = ws['B2']
        logo_cell.value = "관세청 KOREA CUSTOMS SERVICE"
        logo_cell.font = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
        logo_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 메인 타이틀
        ws.merge_cells('B6:N6')
        title_cell = ws['B6']
        title_cell.value = title
        title_cell.font = Font(name='맑은 고딕', size=32, bold=True, color=ColorPalette.PRIMARY)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[6].height = 60
        
        # 서브타이틀
        ws.merge_cells('B8:N8')
        sub_cell = ws['B8']
        sub_cell.value = subtitle
        sub_cell.font = Font(name='맑은 고딕', size=14, color=ColorPalette.SECONDARY)
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 날짜
        ws.merge_cells('B10:N10')
        date_cell = ws['B10']
        date_cell.value = f"보고서 작성일: {self.report_date}"
        date_cell.font = Font(name='맑은 고딕', size=11, color=ColorPalette.DARK_GRAY)
        date_cell.alignment = Alignment(horizontal='center')
        
        # KPI 카드들 (4열)
        row = 14
        col_positions = [2, 5, 8, 11]
        
        for i, (label, value) in enumerate(metrics.items()):
            if i >= 4:
                break
            col = col_positions[i]
            
            # 카드 배경
            for r in range(row, row + 4):
                for c in range(col, col + 3):
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'
                    )
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
            
            # 라벨
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
            label_cell = ws.cell(row=row, column=col)
            label_cell.value = label
            label_cell.font = Font(name='맑은 고딕', size=10, color=ColorPalette.DARK_GRAY)
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 값
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+2)
            value_cell = ws.cell(row=row+1, column=col)
            value_cell.value = value
            value_cell.font = Font(name='맑은 고딕', size=22, bold=True, color=ColorPalette.PRIMARY)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 하단 정보
        ws.merge_cells('B30:N30')
        footer = ws['B30']
        footer.value = "데이터 출처: CLRI_TANSAD_ITM_D, CLRI_TANSAD_UT_PRC_M | 분석 방법론: WCO PMM Framework"
        footer.font = Font(name='맑은 고딕', size=9, color=ColorPalette.DARK_GRAY)
        footer.alignment = Alignment(horizontal='center')
        
        # 열 너비
        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 10
        
        return ws
    
    def _create_executive_summary(self, wb: Workbook) -> Worksheet:
        """경영진 대시보드 시트"""
        ws = wb.create_sheet("경영진 대시보드")
        
        # 요약 데이터
        summary = self.kpi_calc.calc_executive_summary()
        scorecard = self.kpi_calc.calc_kpi_scorecard()
        
        # 제목
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "📊 경영진 대시보드"
        title_cell.font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws.row_dimensions[1].height = 40
        
        # 부제목
        ws.merge_cells('A2:L2')
        ws['A2'].value = f"분석 기간: {summary.get('period', 'N/A')} | 작성일: {self.report_date}"
        ws['A2'].font = Font(name='맑은 고딕', size=10, color=ColorPalette.DARK_GRAY)
        
        # KPI 카드들 (1행에 4개)
        row = 4
        cards = [
            ("총 신고건수", f"{summary['total_declarations']:,.0f}건", "수입신고"),
            ("총 관세수입", format_currency(summary['total_tax_krw'], 'KRW'), "세수실적"),
            ("총 수입금액", format_currency(summary['total_value_usd'], 'USD'), "수입액"),
            ("전년대비 성장률", f"{summary['yoy_growth_pct']:+.1f}%", "YoY"),
        ]
        
        for i, (label, value, sub) in enumerate(cards):
            col = 1 + i * 3
            end_row = add_kpi_card(ws, row, col, label, value, sub)
        
        # KPI 스코어카드 테이블
        scorecard_data = []
        for _, row_data in scorecard.iterrows():
            status_kr = {
                'Excellent': '우수',
                'Good': '양호', 
                'Needs Improvement': '개선필요'
            }.get(row_data['status'], row_data['status'])
            
            scorecard_data.append({
                'name': row_data['name_kr'],
                'actual': row_data['actual'],
                'target': row_data['target'] if pd.notna(row_data['target']) else 0,
                'status': status_kr
            })
        
        add_scorecard_table(ws, 10, 1, scorecard_data, "WCO PMM KPI 스코어카드")
        
        # 리스크 매트릭스
        add_risk_matrix(ws, 10, 8, "리스크 평가 매트릭스")
        
        # 주요 발견사항 섹션
        findings_row = 22
        ws.merge_cells(f'A{findings_row}:L{findings_row}')
        ws[f'A{findings_row}'].value = "📌 주요 발견사항 및 권고사항"
        ws[f'A{findings_row}'].font = Font(name='맑은 고딕', size=14, bold=True, color=ColorPalette.PRIMARY)
        
        findings = [
            f"• 총 {summary['total_declarations']:,.0f}건의 수입신고에서 {format_currency(summary['total_tax_krw'], 'KRW')}의 관세 수입 달성",
            f"• 전년대비 {summary['yoy_growth_pct']:+.1f}%의 {'성장' if summary['yoy_growth_pct'] > 0 else '감소'} 추세",
            f"• 품목 HHI 지수 {summary['hhi_commodity']:.0f}으로 {'보통' if summary['hhi_commodity'] < 1800 else '높은'} 수준의 집중도",
            f"• 최근 과소신고 탐지율 {summary['underval_rate']:.1f}% - 지속적인 모니터링 필요",
            f"• 권고: 고위험 HS코드-국가 조합에 대한 심사 자원 집중 배치"
        ]
        
        for i, finding in enumerate(findings):
            ws[f'A{findings_row + 1 + i}'].value = finding
            ws[f'A{findings_row + 1 + i}'].font = Font(name='맑은 고딕', size=10)
        
        return ws
    
    def _create_methodology_sheet(self, wb: Workbook) -> Worksheet:
        """분석방법론 시트"""
        ws = wb.create_sheet("분석방법론")
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'].value = "📋 데이터 출처 및 분석 방법론"
        ws['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws.row_dimensions[1].height = 40
        
        # 데이터 소스
        ws['A3'].value = "1. 데이터 출처"
        ws['A3'].font = Font(name='맑은 고딕', size=14, bold=True)
        
        sources = [
            ("CLRI_TANSAD_ITM_D", "수입신고 품목 데이터", "약 2,140만 건"),
            ("CLRI_TANSAD_UT_PRC_M", "단가심사 데이터", "약 1,400만 건"),
        ]
        
        for i, (table, desc, size) in enumerate(sources):
            ws[f'A{5+i}'].value = f"  • {table}"
            ws[f'A{5+i}'].font = Font(name='Consolas', size=10)
            ws[f'D{5+i}'].value = desc
            ws[f'G{5+i}'].value = size
        
        # KPI 계산 방법론
        ws['A9'].value = "2. KPI 계산 방법론"
        ws['A9'].font = Font(name='맑은 고딕', size=14, bold=True)
        
        kpi_methods = [
            ("전년대비 성장률 (YoY)", "(당기 세수 - 전기 세수) / 전기 세수 × 100"),
            ("HHI 집중도 지수", "시장점유율 제곱의 합계 × 10,000"),
            ("과소신고율", "심사가격 > 신고가격 × 1.3인 건수 / 전체 건수 × 100"),
            ("변동계수 (CV)", "표준편차 / 평균 × 100"),
        ]
        
        for i, (kpi, formula) in enumerate(kpi_methods):
            ws[f'A{11+i}'].value = f"  • {kpi}:"
            ws[f'A{11+i}'].font = Font(name='맑은 고딕', size=10, bold=True)
            ws[f'D{11+i}'].value = formula
            ws[f'D{11+i}'].font = Font(name='Consolas', size=9)
        
        # 참조 프레임워크
        ws['A17'].value = "3. 참조 프레임워크"
        ws['A17'].font = Font(name='맑은 고딕', size=14, bold=True)
        
        refs = [
            "• WCO 성과측정모델 (PMM) - 4대 차원 (무역원활화, 세수확보, 위험관리, 조직발전)",
            "• WCO 관세 위험관리 개론서 (2022)",
            "• UN Comtrade 데이터베이스 분석 표준",
            "• 관세청 관세연감 통계 방법론",
        ]
        
        for i, ref in enumerate(refs):
            ws[f'A{19+i}'].value = ref
            ws[f'A{19+i}'].font = Font(name='맑은 고딕', size=10)
        
        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['G'].width = 15
        
        return ws
    
    def _create_glossary_sheet(self, wb: Workbook) -> Worksheet:
        """용어정의 시트"""
        ws = wb.create_sheet("용어정의")
        
        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'].value = "📖 용어 정의"
        ws['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws.row_dimensions[1].height = 40
        
        # 헤더
        headers = ['용어 (한글)', '용어 (영문)', '약어', '정의']
        for i, h in enumerate(headers):
            cell = ws.cell(row=3, column=i+1)
            cell.value = h
            cell.font = self.sm.get_header_font()
            cell.fill = self.sm.get_header_fill()
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 용어 목록
        glossary = [
            ("관세", "Customs Duty", "CD", "수입물품에 부과되는 세금"),
            ("품목분류코드", "HS Code", "HS", "국제통일상품분류체계 코드"),
            ("신고가격", "Declared Value", "DV", "수입자가 신고한 물품 가격"),
            ("심사가격", "Assessed Value", "AV", "세관이 결정한 과세 가격"),
            ("과소신고", "Undervaluation", "-", "실제 가격보다 낮게 신고하는 행위"),
            ("허핀달-허쉬만지수", "HHI", "HHI", "시장 집중도 지수 (0~10,000)"),
            ("전년대비", "Year-over-Year", "YoY", "전년 동기 대비 비교"),
            ("전월대비", "Month-over-Month", "MoM", "전월 대비 비교"),
            ("세계관세기구", "WCO", "WCO", "World Customs Organization"),
            ("성과측정모델", "PMM", "PMM", "Performance Measurement Model"),
            ("핵심성과지표", "KPI", "KPI", "Key Performance Indicator"),
            ("리스크 매트릭스", "Risk Matrix", "-", "빈도×심각도 평가 그리드"),
            ("파레토 분석", "Pareto Analysis", "-", "80/20 법칙 기반 분석"),
            ("변동계수", "Coefficient of Variation", "CV", "표준편차/평균 × 100"),
        ]
        
        for i, (kr, en, abbr, defn) in enumerate(glossary):
            row = 4 + i
            ws.cell(row=row, column=1).value = kr
            ws.cell(row=row, column=2).value = en
            ws.cell(row=row, column=3).value = abbr
            ws.cell(row=row, column=4).value = defn
            
            # 스타일
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name='맑은 고딕', size=10)
                cell.border = self.sm.get_thin_border()
                if col == 3:
                    cell.alignment = Alignment(horizontal='center')
        
        # 열 너비
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 45
        
        return ws
    
    # === 관세 수입 현황 보고서 ===
    
    def create_revenue_report(self, output_path: str):
        """프리미엄 관세 수입 현황 보고서 생성"""
        print("📊 프리미엄 관세 수입 현황 보고서 생성 중...")
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # 데이터 조회
        print("  → 데이터 조회...")
        df_yearly = self.kpi_calc.calc_revenue_by_period('yearly')
        df_monthly = self.kpi_calc.calc_revenue_by_period('monthly')
        df_yoy = self.kpi_calc.calc_yoy_growth()
        df_pareto = self.kpi_calc.calc_pareto_analysis('hs2', 'tax')
        hhi_hs = self.kpi_calc.calc_hhi_by_dimension('hs2')
        hhi_country = self.kpi_calc.calc_hhi_by_dimension('country')
        summary = self.kpi_calc.calc_executive_summary()
        
        # 국가별 데이터
        df_country = pd.read_sql("""
            SELECT 
                ORIG_CNTY_CD as country,
                COUNT(*) as declarations,
                SUM(ITM_TAX_AMT) as tax_amount,
                SUM(ITM_INVC_USD_AMT) as value_usd,
                ROUND(SUM(ITM_TAX_AMT) * 100.0 / SUM(SUM(ITM_TAX_AMT)) OVER(), 2) as share_pct
            FROM CLRI_TANSAD_ITM_D
            WHERE DEL_YN = 'N' AND TANSAD_YY >= '23' AND ORIG_CNTY_CD IS NOT NULL
            GROUP BY ORIG_CNTY_CD
            ORDER BY tax_amount DESC
            FETCH FIRST 20 ROWS ONLY
        """, self.conn)
        df_country.columns = ['국가코드', '신고건수', '관세액', '수입금액(USD)', '비중(%)']
        
        # 표지 생성
        print("  → 표지 생성...")
        metrics = {
            '총 신고건수': f"{summary['total_declarations']:,.0f}건",
            '총 관세수입': format_currency(summary['total_tax_krw'], 'KRW'),
            '총 수입금액': format_currency(summary['total_value_usd'], 'USD'),
            '분석 기간': summary['period']
        }
        self._create_cover_sheet(wb, "관세 수입 현황 분석", "Customs Revenue Analysis Report", metrics)
        
        # 경영진 대시보드
        print("  → 경영진 대시보드 생성...")
        self._create_executive_summary(wb)
        
        # 연도별 추이 시트
        print("  → 연도별 추이 시트 생성...")
        ws_yearly = wb.create_sheet("연도별 추이")
        
        # 제목
        ws_yearly.merge_cells('A1:H1')
        ws_yearly['A1'].value = "📈 연도별 관세 수입 추이 분석"
        ws_yearly['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        ws_yearly.row_dimensions[1].height = 35
        
        # 데이터 테이블
        df_yoy_display = df_yoy[['period', 'declaration_count', 'total_tax', 'yoy_growth_pct']].copy()
        df_yoy_display.columns = ['연도', '신고건수', '총세액', '성장률(%)']
        end_row = write_styled_dataframe(ws_yearly, df_yoy_display, start_row=3, title="연도별 관세 수입 및 성장률")
        
        # 콤보 차트 (세액 + 성장률)
        if len(df_yoy) > 1:
            add_combo_chart(
                ws_yearly,
                bar_col=3,  # 총세액
                line_col=4,  # 성장률
                cat_col=1,   # 연도
                start_row=3,
                end_row=3 + len(df_yoy_display),
                position="F3",
                title="관세 수입 및 성장률 추이",
                bar_title="관세 수입",
                line_title="성장률 %"
            )
        
        # 품목별 파레토 분석 시트
        print("  → 파레토 분석 시트 생성...")
        ws_pareto = wb.create_sheet("파레토 분석")
        
        ws_pareto.merge_cells('A1:H1')
        ws_pareto['A1'].value = "📊 파레토 분석 (80/20 법칙)"
        ws_pareto['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        # 파레토 테이블
        pareto_display = df_pareto[['category', 'value', 'share_pct', 'cumulative_pct', 'pareto_zone']].head(20)
        pareto_display.columns = ['HS류', '관세액', '비중(%)', '누적비중(%)', '구간']
        end_row = write_styled_dataframe(ws_pareto, pareto_display, start_row=3, title="HS류별 관세 수입 TOP 20")
        
        # 히트맵 서식 (비중 % 컬럼)
        add_heatmap_formatting(ws_pareto, 6, 5 + len(pareto_display), 3, 3)
        
        # Zone A 카운트 요약
        zone_a = len(df_pareto[df_pareto['pareto_zone'] == 'A (Top 80%)'])
        ws_pareto[f'A{end_row + 2}'].value = f"※ A 구간 (상위 80%): {zone_a}개 HS류가 전체 세수의 80%를 차지"
        ws_pareto[f'A{end_row + 2}'].font = Font(name='맑은 고딕', size=11, bold=True, color=ColorPalette.PRIMARY)
        
        # HHI 요약
        ws_pareto[f'A{end_row + 4}'].value = f"HHI 집중도 지수: {hhi_hs['hhi']:.0f} ({hhi_hs['concentration_level']})"
        ws_pareto[f'A{end_row + 5}'].value = f"상위 5개 품목 비중: {hhi_hs['top_5_share']:.1f}%"
        
        # 국가별 현황 시트
        print("  → 국가별 현황 시트 생성...")
        ws_country = wb.create_sheet("국가별 현황")
        
        ws_country.merge_cells('A1:H1')
        ws_country['A1'].value = "🌍 원산지 국가별 수입 현황"
        ws_country['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        end_row = write_styled_dataframe(ws_country, df_country, start_row=3, title="원산지 국가별 관세 수입 TOP 20")
        
        # 데이터바 (비중 % 컬럼)
        add_databar_formatting(ws_country, 6, 5 + len(df_country), 5)
        
        # HHI 국가
        ws_country[f'A{end_row + 2}'].value = f"HHI 집중도 지수: {hhi_country['hhi']:.0f} ({hhi_country['concentration_level']})"
        
        # 월별 추이 시트
        print("  → 월별 추이 시트 생성...")
        ws_monthly = wb.create_sheet("월별 추이")
        
        ws_monthly.merge_cells('A1:H1')
        ws_monthly['A1'].value = "📅 월별 관세 수입 추이 (최근 36개월)"
        ws_monthly['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        df_monthly_display = df_monthly[['period', 'declaration_count', 'total_tax']].head(36).copy()
        df_monthly_display.columns = ['월', '신고건수', '총세액']
        end_row = write_styled_dataframe(ws_monthly, df_monthly_display, start_row=3, title="월별 관세 수입 데이터")
        
        # 라인 차트
        chart = LineChart()
        chart.style = 10
        chart.title = "월별 관세 수입 추이"
        chart.y_axis.title = "관세액"
        
        data = Reference(ws_monthly, min_col=3, min_row=5, max_row=5 + min(36, len(df_monthly_display)))
        cats = Reference(ws_monthly, min_col=1, min_row=6, max_row=5 + min(36, len(df_monthly_display)))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        if chart.series:
            chart.series[0].graphicalProperties.line.solidFill = ColorPalette.PRIMARY
            chart.series[0].graphicalProperties.line.width = 25000
            chart.series[0].smooth = True
        
        chart.width = 18
        chart.height = 10
        ws_monthly.add_chart(chart, "E3")
        
        # 분석방법론 & 용어정의
        print("  → 부록 시트 생성...")
        self._create_methodology_sheet(wb)
        self._create_glossary_sheet(wb)
        
        # 저장
        wb.save(output_path)
        print(f"✅ 저장 완료: {output_path}")
    
    # === 이상 탐지 보고서 ===
    
    def create_anomaly_report(self, output_path: str):
        """프리미엄 이상 탐지 보고서 생성"""
        print("\n🚨 프리미엄 이상 탐지 보고서 생성 중...")
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # 데이터 조회
        print("  → 데이터 조회...")
        df_underval = self.kpi_calc.calc_undervaluation_stats()
        df_risk = self.kpi_calc.calc_risk_score_by_hs_country()
        df_misclass = self.kpi_calc.calc_hs_misclassification_rate()
        
        # 고위험 업체
        df_importers = pd.read_sql("""
            SELECT 
                IMPPN_TIN as tin,
                MAX(IMPPN_NM) as importer_name,
                COUNT(*) as total,
                SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 
                         AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) as underval,
                ROUND(SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 
                               AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) * 100.0 / COUNT(*), 1) as rate_pct,
                SUM(ASSD_INVC_USD_AMT) as total_value
            FROM CLRI_TANSAD_UT_PRC_M
            WHERE DEL_YN = 'N' AND TANSAD_YY >= '23'
            GROUP BY IMPPN_TIN
            HAVING COUNT(*) >= 50
               AND SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 
                            AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) >= 10
            ORDER BY underval DESC
            FETCH FIRST 30 ROWS ONLY
        """, self.conn)
        df_importers.columns = ['사업자번호', '업체명', '총건수', '과소신고건수', '과소신고율(%)', '총거래액']
        
        # 요약 통계
        total_underval = df_underval['underval_count'].sum() if len(df_underval) > 0 else 0
        total_loss = df_underval['estimated_loss_usd'].sum() if len(df_underval) > 0 else 0
        high_risk_combos = len(df_risk[df_risk['risk_score'] >= 50]) if len(df_risk) > 0 else 0
        
        # 표지
        print("  → 표지 생성...")
        metrics = {
            '과소신고 의심': f"{total_underval:,.0f}건",
            '추정 탈루액': format_currency(total_loss, 'USD'),
            '고위험 조합': f"{high_risk_combos}개",
            '고위험 업체': f"{len(df_importers)}개"
        }
        self._create_cover_sheet(wb, "이상 탐지 분석 리포트", "Anomaly Detection Analysis Report", metrics)
        
        # 리스크 개요 대시보드
        print("  → 리스크 대시보드 생성...")
        ws_exec = wb.create_sheet("리스크 대시보드")
        
        ws_exec.merge_cells('A1:L1')
        ws_exec['A1'].value = "🎯 리스크 평가 대시보드"
        ws_exec['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws_exec.row_dimensions[1].height = 40
        
        # KPI 카드들
        cards = [
            ("과소신고 의심건수", f"{total_underval:,.0f}건", "30% 초과 기준"),
            ("추정 탈루액", format_currency(total_loss, 'USD'), "잠재적 세수손실"),
            ("고위험 조합수", f"{high_risk_combos}개", "리스크점수 50+"),
            ("고위험 업체수", f"{len(df_importers)}개", "반복 위반 업체"),
        ]
        
        for i, (label, value, sub) in enumerate(cards):
            col = 1 + i * 3
            add_kpi_card(ws_exec, 3, col, label, value, sub)
        
        # 리스크 매트릭스
        add_risk_matrix(ws_exec, 9, 1, "리스크 평가 매트릭스")
        
        # 요약 통계 테이블
        underval_summary = df_underval[['period', 'total_count', 'underval_count', 'underval_rate']].head(5)
        underval_summary.columns = ['연도', '총건수', '과소신고건수', '과소신고율(%)']
        write_styled_dataframe(ws_exec, underval_summary, start_row=9, start_col=8, title="연도별 과소신고 추이")
        
        # 과소신고 분석 시트
        print("  → 과소신고 분석 시트 생성...")
        ws_underval = wb.create_sheet("과소신고 분석")
        
        ws_underval.merge_cells('A1:H1')
        ws_underval['A1'].value = "💰 과소신고 탐지 분석"
        ws_underval['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        underval_display = df_underval.copy()
        underval_display.columns = ['연도', '총건수', '과소신고건수', '과소신고율(%)', '추정탈루액(USD)']
        end_row = write_styled_dataframe(ws_underval, underval_display, start_row=3, title="연도별 과소신고 통계")
        
        # 히트맵 (Rate 컬럼)
        add_heatmap_formatting(ws_underval, 6, 5 + len(underval_display), 4, 4, reverse=True)
        
        # 바 차트
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "연도별 과소신고율"
        
        data = Reference(ws_underval, min_col=4, min_row=5, max_row=5 + len(underval_display))
        cats = Reference(ws_underval, min_col=1, min_row=6, max_row=5 + len(underval_display))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = ColorPalette.DANGER
        
        chart.width = 12
        chart.height = 8
        ws_underval.add_chart(chart, "G3")
        
        # HS-국가 리스크 시트
        print("  → HS-국가 리스크 시트 생성...")
        ws_risk = wb.create_sheet("품목국가 리스크")
        
        ws_risk.merge_cells('A1:I1')
        ws_risk['A1'].value = "⚠️ HS코드 × 국가 리스크 분석"
        ws_risk['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        risk_display = df_risk[['hs4', 'country', 'total_count', 'underval_count', 'underval_rate', 'risk_score']].head(30)
        risk_display.columns = ['HS코드', '국가', '총건수', '과소신고건수', '과소신고율(%)', '리스크점수']
        end_row = write_styled_dataframe(ws_risk, risk_display, start_row=3, title="고위험 HS-국가 조합 TOP 30")
        
        # 리스크 점수 히트맵
        add_heatmap_formatting(ws_risk, 6, 5 + len(risk_display), 6, 6, reverse=True)
        
        # 고위험 업체 시트
        print("  → 고위험 업체 시트 생성...")
        ws_importers = wb.create_sheet("고위험 업체")
        
        ws_importers.merge_cells('A1:G1')
        ws_importers['A1'].value = "🏢 고위험 수입업체 분석"
        ws_importers['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        end_row = write_styled_dataframe(ws_importers, df_importers, start_row=3, title="과소신고 다발 업체 TOP 30")
        
        # Rate % 히트맵
        add_heatmap_formatting(ws_importers, 6, 5 + len(df_importers), 5, 5, reverse=True)
        
        # HS 분류 오류 시트
        print("  → HS 분류 오류 시트 생성...")
        ws_misclass = wb.create_sheet("품목분류 오류")
        
        ws_misclass.merge_cells('A1:E1')
        ws_misclass['A1'].value = "🔄 HS코드 분류 오류 분석"
        ws_misclass['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        misclass_display = df_misclass.copy()
        misclass_display.columns = ['연도', '총건수', '분류오류건수', '분류오류율(%)']
        end_row = write_styled_dataframe(ws_misclass, misclass_display, start_row=3, title="연도별 품목분류 오류 통계")
        
        # 부록
        print("  → 부록 시트 생성...")
        self._create_methodology_sheet(wb)
        self._create_glossary_sheet(wb)
        
        # 저장
        wb.save(output_path)
        print(f"✅ 저장 완료: {output_path}")


def main():
    """메인 실행"""
    print("=" * 60)
    print("🚀 프리미엄 관세 분석 보고서 생성 (한국어 버전)")
    print(f"⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # DB 연결
    print("\n🔗 DB 연결 중...")
    try:
        conn = oracledb.connect(**DB_CONFIG)
        print("✅ DB 연결 성공")
    except Exception as e:
        print(f"❌ DB 연결 실패: {e}")
        sys.exit(1)
    
    try:
        generator = PremiumReportGeneratorKR(conn)
        
        # 관세 수입 현황 보고서
        revenue_path = os.path.join(BASE_PATH, "프리미엄_관세수입현황_보고서_KR.xlsx")
        generator.create_revenue_report(revenue_path)
        
        # 이상 탐지 보고서
        anomaly_path = os.path.join(BASE_PATH, "프리미엄_이상탐지_보고서_KR.xlsx")
        generator.create_anomaly_report(anomaly_path)
        
        print("\n" + "=" * 60)
        print("✅ 모든 프리미엄 보고서 생성 완료!")
        print(f"📁 관세 수입 현황: {revenue_path}")
        print(f"📁 이상 탐지: {anomaly_path}")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()
        print("🔌 DB 연결 종료")


if __name__ == "__main__":
    main()
