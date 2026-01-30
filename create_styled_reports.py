#!/usr/bin/env python3
"""
관세 분석 Excel 보고서 스타일링

기존 분석 결과를 보고서 수준의 품질로 업그레이드합니다:
- 표지 디자인
- 헤더 스타일링
- 조건부 서식
- 차트 개선
- 숫자 포맷팅
"""

import oracledb
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment,
    NamedStyle, GradientFill
)
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os

# DB 접속 정보
DB_CONFIG = {
    "user": "CLRIUSR",
    "password": "ntancisclri1!",
    "dsn": "211.239.120.42:3535/NTANCIS"
}

# 색상 팔레트
COLORS = {
    'primary': '1F4E79',      # 진한 파랑
    'secondary': '2E75B6',    # 중간 파랑
    'accent': '5B9BD5',       # 밝은 파랑
    'light': 'D6DCE5',        # 연한 회색
    'white': 'FFFFFF',
    'dark': '333333',
    'success': '70AD47',      # 녹색
    'warning': 'FFC000',      # 노랑
    'danger': 'C00000',       # 빨강
    'orange': 'ED7D31',
}

# 스타일 정의
def create_styles():
    """스타일 생성"""
    styles = {}
    
    # 제목 스타일
    styles['title'] = {
        'font': Font(name='맑은 고딕', size=28, bold=True, color=COLORS['primary']),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }
    
    # 부제목 스타일
    styles['subtitle'] = {
        'font': Font(name='맑은 고딕', size=14, color=COLORS['secondary']),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }
    
    # 헤더 스타일
    styles['header'] = {
        'font': Font(name='맑은 고딕', size=11, bold=True, color=COLORS['white']),
        'fill': PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color=COLORS['white']),
            right=Side(style='thin', color=COLORS['white']),
            top=Side(style='thin', color=COLORS['white']),
            bottom=Side(style='thin', color=COLORS['white'])
        )
    }
    
    # 데이터 스타일
    styles['data'] = {
        'font': Font(name='맑은 고딕', size=10),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin', color=COLORS['light']),
            right=Side(style='thin', color=COLORS['light']),
            top=Side(style='thin', color=COLORS['light']),
            bottom=Side(style='thin', color=COLORS['light'])
        )
    }
    
    # 숫자 스타일
    styles['number'] = {
        'font': Font(name='맑은 고딕', size=10),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin', color=COLORS['light']),
            right=Side(style='thin', color=COLORS['light']),
            top=Side(style='thin', color=COLORS['light']),
            bottom=Side(style='thin', color=COLORS['light'])
        )
    }
    
    # KPI 카드 스타일
    styles['kpi_label'] = {
        'font': Font(name='맑은 고딕', size=10, color=COLORS['dark']),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }
    
    styles['kpi_value'] = {
        'font': Font(name='맑은 고딕', size=24, bold=True, color=COLORS['primary']),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }
    
    return styles


def apply_style(cell, style_dict):
    """셀에 스타일 적용"""
    for key, value in style_dict.items():
        setattr(cell, key, value)


def create_cover_sheet(wb, title, subtitle, metrics):
    """표지 시트 생성"""
    ws = wb.create_sheet("표지", 0)
    styles = create_styles()
    
    # 배경색 설정
    for row in range(1, 50):
        for col in range(1, 20):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=COLORS['white'], 
                end_color=COLORS['white'], 
                fill_type='solid'
            )
    
    # 상단 배너
    for col in range(1, 15):
        ws.cell(row=1, column=col).fill = PatternFill(
            start_color=COLORS['primary'], 
            end_color=COLORS['primary'], 
            fill_type='solid'
        )
        ws.cell(row=2, column=col).fill = PatternFill(
            start_color=COLORS['primary'], 
            end_color=COLORS['primary'], 
            fill_type='solid'
        )
    
    # 제목
    ws.merge_cells('B5:M5')
    title_cell = ws['B5']
    title_cell.value = title
    apply_style(title_cell, styles['title'])
    ws.row_dimensions[5].height = 50
    
    # 부제목
    ws.merge_cells('B7:M7')
    subtitle_cell = ws['B7']
    subtitle_cell.value = subtitle
    apply_style(subtitle_cell, styles['subtitle'])
    
    # 생성 일시
    ws.merge_cells('B9:M9')
    date_cell = ws['B9']
    date_cell.value = f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}"
    date_cell.font = Font(name='맑은 고딕', size=11, color=COLORS['secondary'])
    date_cell.alignment = Alignment(horizontal='center')
    
    # KPI 카드들
    row = 12
    col_positions = [2, 5, 8, 11]
    
    for i, (label, value) in enumerate(metrics.items()):
        if i >= 4:
            break
        col = col_positions[i]
        
        # 카드 배경
        for r in range(row, row + 4):
            for c in range(col, col + 3):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color=COLORS['light'], 
                    end_color=COLORS['light'], 
                    fill_type='solid'
                )
        
        # 라벨
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = label
        apply_style(label_cell, styles['kpi_label'])
        
        # 값
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+2)
        value_cell = ws.cell(row=row+1, column=col)
        value_cell.value = value
        apply_style(value_cell, styles['kpi_value'])
    
    # 열 너비 조정
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    return ws


def style_data_sheet(ws, df, sheet_title, has_chart=False):
    """데이터 시트 스타일링"""
    styles = create_styles()
    
    # 시트 제목 (A1:전체 컬럼 병합)
    num_cols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"📊 {sheet_title}"
    title_cell.font = Font(name='맑은 고딕', size=16, bold=True, color=COLORS['primary'])
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # 빈 행
    ws.row_dimensions[2].height = 10
    
    # 데이터 시작 행
    start_row = 3
    
    # 헤더 작성
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = col_name
        apply_style(cell, styles['header'])
    ws.row_dimensions[start_row].height = 30
    
    # 데이터 작성
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        # 줄무늬 배경
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            row_fill = PatternFill(start_color=COLORS['white'], end_color=COLORS['white'], fill_type='solid')
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = row_fill
            
            # 숫자 포맷팅
            if isinstance(value, (int, float)):
                if value > 1000000000:
                    cell.number_format = '#,##0,,"B"'
                elif value > 1000000:
                    cell.number_format = '#,##0,,"M"'
                else:
                    cell.number_format = '#,##0'
                apply_style(cell, styles['number'])
            else:
                apply_style(cell, styles['data'])
        
        ws.row_dimensions[row_idx].height = 22
    
    # 열 너비 자동 조정
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(start_row + 1, start_row + 1 + len(df)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 4, 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    return ws


def add_bar_chart(ws, title, data_range, cat_range, position, width=15, height=10):
    """막대 차트 추가"""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.legend = None
    
    data = Reference(ws, **data_range)
    cats = Reference(ws, **cat_range)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # 색상 설정
    chart.series[0].graphicalProperties.solidFill = COLORS['accent']
    
    chart.width = width
    chart.height = height
    
    ws.add_chart(chart, position)
    return chart


def add_line_chart(ws, title, data_range, cat_range, position, width=18, height=10):
    """라인 차트 추가"""
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    
    data = Reference(ws, **data_range)
    cats = Reference(ws, **cat_range)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    chart.series[0].graphicalProperties.line.solidFill = COLORS['primary']
    chart.series[0].graphicalProperties.line.width = 25000  # EMUs
    chart.series[0].smooth = True
    
    chart.width = width
    chart.height = height
    
    ws.add_chart(chart, position)
    return chart


def add_doughnut_chart(ws, title, data_range, cat_range, position, width=12, height=10):
    """도넛 차트 추가"""
    chart = DoughnutChart()
    chart.title = title
    chart.style = 10
    
    data = Reference(ws, **data_range)
    cats = Reference(ws, **cat_range)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    chart.width = width
    chart.height = height
    
    ws.add_chart(chart, position)
    return chart


def add_conditional_formatting(ws, start_row, end_row, col, rule_type='colorscale'):
    """조건부 서식 추가"""
    col_letter = get_column_letter(col)
    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"
    
    if rule_type == 'colorscale':
        rule = ColorScaleRule(
            start_type='min', start_color='63BE7B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='F8696B'
        )
        ws.conditional_formatting.add(range_str, rule)
    elif rule_type == 'databar':
        rule = DataBarRule(
            start_type='min', start_value=0,
            end_type='max', end_value=100,
            color=COLORS['accent']
        )
        ws.conditional_formatting.add(range_str, rule)


def create_revenue_report(conn, output_path):
    """관세 수입 현황 보고서 생성"""
    print("📊 관세 수입 현황 보고서 생성 중...")
    
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 데이터 조회
    print("  → 연도별 데이터 조회...")
    df_yearly = pd.read_sql("""
        SELECT 
            '20' || TANSAD_YY as 연도,
            COUNT(*) as 건수,
            SUM(ITM_TAX_AMT) as 총세액,
            SUM(ITM_INVC_USD_AMT) as "총수입액(USD)"
        FROM CLRI_TANSAD_ITM_D
        WHERE DEL_YN = 'N' AND TANSAD_YY >= '20'
        GROUP BY TANSAD_YY
        ORDER BY TANSAD_YY DESC
    """, conn)
    
    print("  → 품목별 데이터 조회...")
    df_commodity = pd.read_sql("""
        SELECT 
            SUBSTR(HS_CD, 1, 2) as "HS코드",
            COUNT(*) as 건수,
            SUM(ITM_TAX_AMT) as 총세액,
            ROUND(SUM(ITM_TAX_AMT) * 100.0 / SUM(SUM(ITM_TAX_AMT)) OVER(), 1) as "비중(%)"
        FROM CLRI_TANSAD_ITM_D
        WHERE DEL_YN = 'N' AND ITM_TAX_AMT > 0
        GROUP BY SUBSTR(HS_CD, 1, 2)
        ORDER BY 총세액 DESC
        FETCH FIRST 15 ROWS ONLY
    """, conn)
    
    print("  → 국가별 데이터 조회...")
    df_country = pd.read_sql("""
        SELECT 
            ORIG_CNTY_CD as "국가코드",
            COUNT(*) as 건수,
            SUM(ITM_TAX_AMT) as 총세액,
            SUM(ITM_INVC_USD_AMT) as "총수입액(USD)"
        FROM CLRI_TANSAD_ITM_D
        WHERE DEL_YN = 'N' AND ORIG_CNTY_CD IS NOT NULL
        GROUP BY ORIG_CNTY_CD
        ORDER BY "총수입액(USD)" DESC NULLS LAST
        FETCH FIRST 15 ROWS ONLY
    """, conn)
    
    print("  → 월별 데이터 조회...")
    df_monthly = pd.read_sql("""
        SELECT 
            TO_CHAR(FRST_RGSR_DTM, 'YYYY-MM') as 월,
            COUNT(*) as 건수,
            SUM(ITM_TAX_AMT) as 총세액
        FROM CLRI_TANSAD_ITM_D
        WHERE DEL_YN = 'N' 
          AND FRST_RGSR_DTM >= ADD_MONTHS(SYSDATE, -24)
        GROUP BY TO_CHAR(FRST_RGSR_DTM, 'YYYY-MM')
        ORDER BY 월
    """, conn)
    
    # 표지 생성
    metrics = {
        '총 건수': f"{df_yearly['건수'].sum():,.0f}",
        '총 세액': f"{df_yearly['총세액'].sum()/1e12:.1f}조",
        '총 수입액': f"${df_yearly['총수입액(USD)'].sum()/1e9:.0f}B",
        '분석 기간': f"{df_yearly['연도'].min()}~{df_yearly['연도'].max()}"
    }
    create_cover_sheet(wb, "관세 수입 현황 분석", "Customs Revenue Analysis Report", metrics)
    
    # 연도별 추이 시트
    print("  → 연도별 추이 시트 생성...")
    ws_yearly = wb.create_sheet("연도별 추이")
    style_data_sheet(ws_yearly, df_yearly, "연도별 관세 수입 추이")
    
    # 차트 추가
    add_bar_chart(
        ws_yearly, "연도별 총세액",
        {'min_col': 3, 'min_row': 3, 'max_row': 3 + len(df_yearly), 'max_col': 3},
        {'min_col': 1, 'min_row': 4, 'max_row': 3 + len(df_yearly)},
        "F3"
    )
    
    # 품목별 현황 시트
    print("  → 품목별 현황 시트 생성...")
    ws_commodity = wb.create_sheet("품목별 현황")
    style_data_sheet(ws_commodity, df_commodity, "HS코드별 관세 수입 TOP 15")
    
    # 조건부 서식 (비중 컬럼)
    add_conditional_formatting(ws_commodity, 4, 4 + len(df_commodity) - 1, 4, 'databar')
    
    # 도넛 차트
    add_doughnut_chart(
        ws_commodity, "품목별 세액 비중",
        {'min_col': 3, 'min_row': 3, 'max_row': 3 + min(10, len(df_commodity)), 'max_col': 3},
        {'min_col': 1, 'min_row': 4, 'max_row': 3 + min(10, len(df_commodity))},
        "F3"
    )
    
    # 국가별 현황 시트
    print("  → 국가별 현황 시트 생성...")
    ws_country = wb.create_sheet("국가별 현황")
    style_data_sheet(ws_country, df_country, "원산지 국가별 수입 현황 TOP 15")
    
    add_bar_chart(
        ws_country, "국가별 수입액",
        {'min_col': 4, 'min_row': 3, 'max_row': 3 + len(df_country), 'max_col': 4},
        {'min_col': 1, 'min_row': 4, 'max_row': 3 + len(df_country)},
        "F3"
    )
    
    # 월별 추이 시트
    print("  → 월별 추이 시트 생성...")
    ws_monthly = wb.create_sheet("월별 추이")
    style_data_sheet(ws_monthly, df_monthly, "최근 24개월 관세 수입 추이")
    
    add_line_chart(
        ws_monthly, "월별 세액 추이",
        {'min_col': 3, 'min_row': 3, 'max_row': 3 + len(df_monthly), 'max_col': 3},
        {'min_col': 1, 'min_row': 4, 'max_row': 3 + len(df_monthly)},
        "E3"
    )
    
    # 저장
    wb.save(output_path)
    print(f"✅ 저장 완료: {output_path}")


def create_anomaly_report(conn, output_path):
    """이상 탐지 보고서 생성"""
    print("\n🚨 이상 탐지 보고서 생성 중...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 데이터 조회
    print("  → 과소신고 데이터 조회...")
    df_underval = pd.read_sql("""
        SELECT 
            SUBSTR(ASSD_HS_CD, 1, 4) as "HS코드",
            ORIG_CNTY_CD as "국가",
            COUNT(*) as 건수,
            ROUND(AVG((ASSD_UT_USD_VAL - DCLD_UT_USD_VAL) / NULLIF(DCLD_UT_USD_VAL, 0) * 100), 1) as "평균차이(%)",
            SUM(ASSD_INVC_USD_AMT - DCLD_INVC_USD_AMT) as "총차액(USD)"
        FROM CLRI_TANSAD_UT_PRC_M
        WHERE DEL_YN = 'N'
          AND DCLD_UT_USD_VAL > 0
          AND ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3
          AND TANSAD_YY >= '23'
        GROUP BY SUBSTR(ASSD_HS_CD, 1, 4), ORIG_CNTY_CD
        HAVING COUNT(*) >= 10
        ORDER BY "총차액(USD)" DESC NULLS LAST
        FETCH FIRST 30 ROWS ONLY
    """, conn)
    
    print("  → 리스크 분석 데이터 조회...")
    df_risk = pd.read_sql("""
        WITH risk_data AS (
            SELECT 
                SUBSTR(ASSD_HS_CD, 1, 4) as HS4,
                ORIG_CNTY_CD,
                CASE WHEN DCLD_HS_CD != ASSD_HS_CD THEN 1 ELSE 0 END as HS_CHANGED,
                CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.5 AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END as UNDERVALUED
            FROM CLRI_TANSAD_UT_PRC_M
            WHERE DEL_YN = 'N' AND TANSAD_YY >= '23'
        )
        SELECT 
            HS4 as "HS코드",
            ORIG_CNTY_CD as "국가",
            COUNT(*) as "총건수",
            SUM(UNDERVALUED) as "과소신고",
            ROUND(SUM(UNDERVALUED) * 100.0 / COUNT(*), 1) as "과소신고율(%)",
            ROUND(SUM(UNDERVALUED) * 3.0 / COUNT(*) * 100 + SUM(HS_CHANGED) * 2.0 / COUNT(*) * 100, 1) as "리스크점수"
        FROM risk_data
        WHERE HS4 IS NOT NULL
        GROUP BY HS4, ORIG_CNTY_CD
        HAVING SUM(UNDERVALUED) >= 50
        ORDER BY "리스크점수" DESC
        FETCH FIRST 30 ROWS ONLY
    """, conn)
    
    print("  → 고위험 업체 데이터 조회...")
    df_importers = pd.read_sql("""
        SELECT 
            IMPPN_TIN as "사업자번호",
            MAX(IMPPN_NM) as "업체명",
            COUNT(*) as "총건수",
            SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) as "과소신고건수",
            ROUND(SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) * 100.0 / COUNT(*), 1) as "과소신고율(%)",
            SUM(ASSD_INVC_USD_AMT) as "총거래액(USD)"
        FROM CLRI_TANSAD_UT_PRC_M
        WHERE DEL_YN = 'N' AND TANSAD_YY >= '23'
        GROUP BY IMPPN_TIN
        HAVING COUNT(*) >= 20
           AND SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) >= 5
        ORDER BY "과소신고건수" DESC
        FETCH FIRST 30 ROWS ONLY
    """, conn)
    
    # 표지 생성
    total_underval = df_underval['건수'].sum() if len(df_underval) > 0 else 0
    total_diff = df_underval['총차액(USD)'].sum() if len(df_underval) > 0 else 0
    high_risk = len(df_risk[df_risk['리스크점수'] >= 80]) if len(df_risk) > 0 else 0
    
    metrics = {
        '과소신고 의심': f"{total_underval:,.0f}건",
        '추정 탈루액': f"${total_diff/1e9:.1f}B",
        '고위험 조합': f"{high_risk}개",
        '고위험 업체': f"{len(df_importers)}개"
    }
    create_cover_sheet(wb, "관세 이상 탐지 리포트", "Customs Anomaly Detection Report", metrics)
    
    # 과소신고 의심 시트
    print("  → 과소신고 시트 생성...")
    ws_underval = wb.create_sheet("과소신고 의심")
    style_data_sheet(ws_underval, df_underval, "과소신고 의심 건 TOP 30")
    
    # 조건부 서식
    if len(df_underval) > 0:
        add_conditional_formatting(ws_underval, 4, 4 + len(df_underval) - 1, 4, 'colorscale')
    
    # 리스크 분석 시트
    print("  → 리스크 분석 시트 생성...")
    ws_risk = wb.create_sheet("품목국가 리스크")
    style_data_sheet(ws_risk, df_risk, "품목-국가 리스크 분석 TOP 30")
    
    if len(df_risk) > 0:
        add_conditional_formatting(ws_risk, 4, 4 + len(df_risk) - 1, 6, 'colorscale')
        
        add_bar_chart(
            ws_risk, "리스크 점수 분포",
            {'min_col': 6, 'min_row': 3, 'max_row': 3 + min(15, len(df_risk)), 'max_col': 6},
            {'min_col': 1, 'min_row': 4, 'max_row': 3 + min(15, len(df_risk))},
            "H3"
        )
    
    # 고위험 업체 시트
    print("  → 고위험 업체 시트 생성...")
    ws_importers = wb.create_sheet("고위험 업체")
    style_data_sheet(ws_importers, df_importers, "고위험 업체 TOP 30")
    
    if len(df_importers) > 0:
        add_conditional_formatting(ws_importers, 4, 4 + len(df_importers) - 1, 5, 'colorscale')
    
    # 저장
    wb.save(output_path)
    print(f"✅ 저장 완료: {output_path}")


def main():
    print("🚀 관세 분석 보고서 생성 시작")
    print(f"⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)
    
    # DB 연결
    print("\n🔗 DB 연결 중...")
    conn = oracledb.connect(**DB_CONFIG)
    print("✅ DB 연결 성공")
    
    try:
        # 출력 경로
        base_path = os.path.dirname(os.path.abspath(__file__))
        
        # 관세 수입 현황 보고서
        revenue_path = os.path.join(base_path, "관세수입현황_분석보고서.xlsx")
        create_revenue_report(conn, revenue_path)
        
        # 이상 탐지 보고서
        anomaly_path = os.path.join(base_path, "관세이상탐지_분석보고서.xlsx")
        create_anomaly_report(conn, anomaly_path)
        
        print("\n" + "=" * 50)
        print("✅ 모든 보고서 생성 완료!")
        print(f"📁 관세 수입 현황: {revenue_path}")
        print(f"📁 이상 탐지: {anomaly_path}")
        
    finally:
        conn.close()
        print("🔌 DB 연결 종료")


if __name__ == "__main__":
    main()
