#!/usr/bin/env python3
"""
프리미엄 관세 보고서 시각화 모듈

WCO/KCS/UN Comtrade 스타일 고급 차트:
- Heatmap (품목×국가, 월×품목)
- Pareto Chart (80/20 분석)
- Risk Matrix (빈도×심각도)
- Combo Charts (세액+성장률)
- Gauge Charts (목표 대비 실적)
- Sparklines & Trend Indicators

References:
- UN Comtrade Analytical Tables
- WCO Annual Report Visualizations
- KCS 관세연감 디자인
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.chart import (
    BarChart, LineChart, PieChart, DoughnutChart, 
    AreaChart, ScatterChart, Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule, FormulaRule
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple, Optional, Any
import colorsys


# === 색상 팔레트 (WCO/UN Comtrade 스타일) ===

class ColorPalette:
    """전문 보고서 색상 팔레트"""
    
    # Primary Colors (UN Blue inspired)
    PRIMARY = '1F4E79'       # 진한 파랑 (헤더)
    SECONDARY = '2E75B6'     # 중간 파랑
    ACCENT = '5B9BD5'        # 밝은 파랑
    
    # Background Colors
    WHITE = 'FFFFFF'
    LIGHT_GRAY = 'F2F2F2'
    MEDIUM_GRAY = 'D6DCE5'
    DARK_GRAY = '333333'
    
    # Status Colors
    SUCCESS = '70AD47'       # 녹색 (Good)
    WARNING = 'FFC000'       # 노랑 (Caution)
    DANGER = 'C00000'        # 빨강 (Alert)
    INFO = '5B9BD5'          # 파랑 (Info)
    
    # Heatmap Gradient
    HEAT_LOW = '63BE7B'      # 녹색 (낮음)
    HEAT_MID = 'FFEB84'      # 노랑 (중간)
    HEAT_HIGH = 'F8696B'     # 빨강 (높음)
    
    # Chart Series Colors (최대 10개)
    SERIES = [
        '1F4E79', '2E75B6', '5B9BD5', '70AD47', 'FFC000',
        'C00000', 'ED7D31', '7030A0', '00B0F0', '00B050'
    ]
    
    @classmethod
    def get_gradient(cls, n: int, start_color: str = None, end_color: str = None) -> List[str]:
        """n개의 그라데이션 색상 생성"""
        if start_color is None:
            start_color = cls.PRIMARY
        if end_color is None:
            end_color = cls.ACCENT
        
        # HEX to RGB
        def hex_to_rgb(hex_color: str) -> Tuple[int, int, int]:
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        
        def rgb_to_hex(rgb: Tuple[int, int, int]) -> str:
            return '{:02X}{:02X}{:02X}'.format(*rgb)
        
        start_rgb = hex_to_rgb(start_color)
        end_rgb = hex_to_rgb(end_color)
        
        colors = []
        for i in range(n):
            ratio = i / max(n - 1, 1)
            rgb = tuple(int(start_rgb[j] + (end_rgb[j] - start_rgb[j]) * ratio) for j in range(3))
            colors.append(rgb_to_hex(rgb))
        
        return colors


class StyleManager:
    """스타일 관리자"""
    
    def __init__(self):
        self.palette = ColorPalette()
        self._styles_created = False
    
    def get_header_fill(self) -> PatternFill:
        return PatternFill(start_color=self.palette.PRIMARY, end_color=self.palette.PRIMARY, fill_type='solid')
    
    def get_subheader_fill(self) -> PatternFill:
        return PatternFill(start_color=self.palette.SECONDARY, end_color=self.palette.SECONDARY, fill_type='solid')
    
    def get_data_fill(self, alt: bool = False) -> PatternFill:
        color = self.palette.LIGHT_GRAY if alt else self.palette.WHITE
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    def get_status_fill(self, status: str) -> PatternFill:
        colors = {
            'excellent': self.palette.SUCCESS,
            'good': self.palette.INFO,
            'warning': self.palette.WARNING,
            'danger': self.palette.DANGER,
            'needs improvement': self.palette.WARNING,
        }
        color = colors.get(status.lower(), self.palette.MEDIUM_GRAY)
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    def get_header_font(self) -> Font:
        return Font(name='맑은 고딕', size=11, bold=True, color=self.palette.WHITE)
    
    def get_title_font(self) -> Font:
        return Font(name='맑은 고딕', size=16, bold=True, color=self.palette.PRIMARY)
    
    def get_data_font(self) -> Font:
        return Font(name='맑은 고딕', size=10)
    
    def get_thin_border(self) -> Border:
        side = Side(style='thin', color=self.palette.MEDIUM_GRAY)
        return Border(left=side, right=side, top=side, bottom=side)
    
    def get_center_alignment(self, wrap: bool = False) -> Alignment:
        return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


# === 고급 차트 함수 ===

def add_pareto_chart(
    ws: Worksheet,
    data_col: int,
    cat_col: int,
    start_row: int,
    end_row: int,
    position: str,
    title: str = "Pareto Analysis",
    width: int = 18,
    height: int = 12
) -> None:
    """파레토 차트 (막대 + 누적선)
    
    80/20 분석을 위한 콤보 차트
    - 막대: 개별 값 (내림차순)
    - 선: 누적 비율 (%)
    """
    # 막대 차트 (주축)
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = title
    
    bar_data = Reference(ws, min_col=data_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.series[0].graphicalProperties.solidFill = ColorPalette.ACCENT
    
    # 라인 차트 (보조축) - 누적 비율
    line_chart = LineChart()
    line_chart.style = 10
    
    # 누적 비율 컬럼이 data_col + 1에 있다고 가정
    cumul_col = data_col + 1
    line_data = Reference(ws, min_col=cumul_col, min_row=start_row, max_row=end_row)
    line_chart.add_data(line_data, titles_from_data=True)
    
    # 라인 스타일
    if line_chart.series:
        line_chart.series[0].graphicalProperties.line.solidFill = ColorPalette.DANGER
        line_chart.series[0].graphicalProperties.line.width = 25000
        line_chart.series[0].smooth = False
        # 마커 추가
        line_chart.series[0].marker = Marker(symbol='circle', size=5)
        line_chart.series[0].marker.graphicalProperties.solidFill = ColorPalette.DANGER
    
    # 보조 Y축
    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"
    
    # 차트 합치기
    bar_chart += line_chart
    bar_chart.y_axis.title = "Value"
    bar_chart.y_axis.axId = 100
    bar_chart.legend.position = 'b'
    
    bar_chart.width = width
    bar_chart.height = height
    
    ws.add_chart(bar_chart, position)


def add_combo_chart(
    ws: Worksheet,
    bar_col: int,
    line_col: int,
    cat_col: int,
    start_row: int,
    end_row: int,
    position: str,
    title: str = "Combo Chart",
    bar_title: str = "Value",
    line_title: str = "Growth %",
    width: int = 18,
    height: int = 12
) -> None:
    """콤보 차트 (막대 + 선)
    
    세액 + 성장률 등 이중 축 표현
    """
    # 막대 차트
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = title
    
    bar_data = Reference(ws, min_col=bar_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.series[0].graphicalProperties.solidFill = ColorPalette.PRIMARY
    
    # 라인 차트
    line_chart = LineChart()
    line_data = Reference(ws, min_col=line_col, min_row=start_row, max_row=end_row)
    line_chart.add_data(line_data, titles_from_data=True)
    
    if line_chart.series:
        line_chart.series[0].graphicalProperties.line.solidFill = ColorPalette.DANGER
        line_chart.series[0].graphicalProperties.line.width = 25000
        line_chart.series[0].smooth = True
    
    # 보조축
    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"
    line_chart.y_axis.title = line_title
    
    # 합치기
    bar_chart += line_chart
    bar_chart.y_axis.title = bar_title
    bar_chart.legend.position = 'b'
    
    bar_chart.width = width
    bar_chart.height = height
    
    ws.add_chart(bar_chart, position)


def add_risk_matrix(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    title: str = "Risk Matrix"
) -> int:
    """리스크 매트릭스 (5x5 그리드)
    
    Frequency(빈도) × Severity(심각도) 매트릭스
    Returns: 종료 행 번호
    """
    sm = StyleManager()
    
    # 제목
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = title
    title_cell.font = sm.get_title_font()
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 5
    )
    
    # 매트릭스 헤더 (Severity)
    severity_labels = ['Very Low', 'Low', 'Medium', 'High', 'Very High']
    freq_labels = ['Very High', 'High', 'Medium', 'Low', 'Very Low']
    
    # Severity 헤더 (상단)
    for i, label in enumerate(severity_labels):
        cell = ws.cell(row=start_row + 2, column=start_col + 1 + i)
        cell.value = label
        cell.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
        cell.fill = sm.get_header_fill()
        cell.alignment = sm.get_center_alignment()
    
    # Frequency 헤더 (좌측) + 매트릭스 셀
    risk_colors = [
        # Severity: VL, L, M, H, VH
        ['70AD47', '70AD47', 'FFC000', 'FFC000', 'C00000'],  # Freq VH
        ['70AD47', 'FFC000', 'FFC000', 'C00000', 'C00000'],  # Freq H
        ['70AD47', 'FFC000', 'FFC000', 'C00000', 'C00000'],  # Freq M
        ['70AD47', '70AD47', 'FFC000', 'FFC000', 'C00000'],  # Freq L
        ['70AD47', '70AD47', '70AD47', 'FFC000', 'FFC000'],  # Freq VL
    ]
    
    risk_labels = [
        ['Low', 'Low', 'Medium', 'Medium', 'High'],
        ['Low', 'Medium', 'Medium', 'High', 'High'],
        ['Low', 'Medium', 'Medium', 'High', 'High'],
        ['Low', 'Low', 'Medium', 'Medium', 'High'],
        ['Low', 'Low', 'Low', 'Medium', 'Medium'],
    ]
    
    for row_idx, freq_label in enumerate(freq_labels):
        # Frequency 헤더
        freq_cell = ws.cell(row=start_row + 3 + row_idx, column=start_col)
        freq_cell.value = freq_label
        freq_cell.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
        freq_cell.fill = sm.get_header_fill()
        freq_cell.alignment = sm.get_center_alignment()
        
        # 매트릭스 셀
        for col_idx in range(5):
            cell = ws.cell(row=start_row + 3 + row_idx, column=start_col + 1 + col_idx)
            cell.value = risk_labels[row_idx][col_idx]
            cell.fill = PatternFill(
                start_color=risk_colors[row_idx][col_idx],
                end_color=risk_colors[row_idx][col_idx],
                fill_type='solid'
            )
            cell.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
            cell.alignment = sm.get_center_alignment()
    
    # 축 라벨
    ws.cell(row=start_row + 1, column=start_col + 3).value = "Severity →"
    ws.cell(row=start_row + 1, column=start_col + 3).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=start_row + 1, column=start_col + 3).alignment = Alignment(horizontal='center')
    
    # 열 너비
    for col in range(start_col, start_col + 6):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    return start_row + 8


def add_heatmap_formatting(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    reverse: bool = False
) -> None:
    """셀 범위에 히트맵 색상 스케일 적용
    
    Args:
        reverse: True면 낮은값=빨강, 높은값=녹색 (비용/리스크)
    """
    if reverse:
        start_color = ColorPalette.HEAT_HIGH
        mid_color = ColorPalette.HEAT_MID
        end_color = ColorPalette.HEAT_LOW
    else:
        start_color = ColorPalette.HEAT_LOW
        mid_color = ColorPalette.HEAT_MID
        end_color = ColorPalette.HEAT_HIGH
    
    range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    
    rule = ColorScaleRule(
        start_type='min', start_color=start_color,
        mid_type='percentile', mid_value=50, mid_color=mid_color,
        end_type='max', end_color=end_color
    )
    
    ws.conditional_formatting.add(range_str, rule)


def add_databar_formatting(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    col: int,
    color: str = None
) -> None:
    """데이터바 조건부 서식"""
    if color is None:
        color = ColorPalette.ACCENT
    
    col_letter = get_column_letter(col)
    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"
    
    rule = DataBarRule(
        start_type='min', start_value=0,
        end_type='max', end_value=100,
        color=color
    )
    
    ws.conditional_formatting.add(range_str, rule)


def add_gauge_indicator(
    ws: Worksheet,
    row: int,
    col: int,
    actual: float,
    target: float,
    min_val: float = 0,
    max_val: float = 100,
    label: str = ""
) -> None:
    """게이지 인디케이터 (목표 대비 실적)
    
    텍스트 기반 간단 게이지: [████████░░] 80%
    """
    # 진행률 계산
    if max_val - min_val > 0:
        progress = (actual - min_val) / (max_val - min_val)
    else:
        progress = 0
    progress = max(0, min(1, progress))
    
    # 10칸 게이지
    filled = int(progress * 10)
    gauge_bar = '█' * filled + '░' * (10 - filled)
    
    # 상태 색상
    if actual >= target:
        status_color = ColorPalette.SUCCESS
    elif actual >= target * 0.8:
        status_color = ColorPalette.WARNING
    else:
        status_color = ColorPalette.DANGER
    
    # 라벨
    if label:
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = label
        label_cell.font = Font(name='맑은 고딕', size=10)
        col += 1
    
    # 게이지 바
    gauge_cell = ws.cell(row=row, column=col)
    gauge_cell.value = f"[{gauge_bar}]"
    gauge_cell.font = Font(name='Consolas', size=10, color=status_color)
    
    # 값
    value_cell = ws.cell(row=row, column=col + 1)
    value_cell.value = f"{actual:.1f} / {target:.1f}"
    value_cell.font = Font(name='맑은 고딕', size=10, bold=True, color=status_color)


def get_trend_arrow(current: float, previous: float, threshold: float = 0.05) -> Tuple[str, str]:
    """추세 화살표 반환
    
    Returns:
        (arrow_symbol, color)
    """
    if previous == 0:
        return ('→', ColorPalette.DARK_GRAY)
    
    change = (current - previous) / previous
    
    if change > threshold:
        return ('↑', ColorPalette.SUCCESS)
    elif change < -threshold:
        return ('↓', ColorPalette.DANGER)
    else:
        return ('→', ColorPalette.DARK_GRAY)


def add_kpi_card(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    title: str,
    value: str,
    subtitle: str = "",
    trend: str = None,
    trend_value: str = None
) -> int:
    """KPI 카드 (대시보드용)
    
    ┌─────────────────┐
    │     TITLE       │
    │   ███ VALUE ███ │
    │   subtitle ↑+5% │
    └─────────────────┘
    
    Returns: 종료 행 번호
    """
    sm = StyleManager()
    
    # 카드 배경 (3행 × 3열)
    card_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    border = Border(
        left=Side(style='thin', color=ColorPalette.MEDIUM_GRAY),
        right=Side(style='thin', color=ColorPalette.MEDIUM_GRAY),
        top=Side(style='thin', color=ColorPalette.MEDIUM_GRAY),
        bottom=Side(style='thin', color=ColorPalette.MEDIUM_GRAY)
    )
    
    for r in range(start_row, start_row + 4):
        for c in range(start_col, start_col + 3):
            cell = ws.cell(row=r, column=c)
            cell.fill = card_fill
            cell.border = border
    
    # 제목
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 2
    )
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = title
    title_cell.font = Font(name='맑은 고딕', size=10, color=ColorPalette.DARK_GRAY)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 값 (큰 숫자)
    ws.merge_cells(
        start_row=start_row + 1, start_column=start_col,
        end_row=start_row + 2, end_column=start_col + 2
    )
    value_cell = ws.cell(row=start_row + 1, column=start_col)
    value_cell.value = value
    value_cell.font = Font(name='맑은 고딕', size=24, bold=True, color=ColorPalette.PRIMARY)
    value_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 부제목 + 트렌드
    ws.merge_cells(
        start_row=start_row + 3, start_column=start_col,
        end_row=start_row + 3, end_column=start_col + 2
    )
    sub_cell = ws.cell(row=start_row + 3, column=start_col)
    
    if trend and trend_value:
        sub_text = f"{subtitle}  {trend}{trend_value}"
    else:
        sub_text = subtitle
    
    sub_cell.value = sub_text
    
    # 트렌드 색상
    if trend == '↑':
        trend_color = ColorPalette.SUCCESS
    elif trend == '↓':
        trend_color = ColorPalette.DANGER
    else:
        trend_color = ColorPalette.DARK_GRAY
    
    sub_cell.font = Font(name='맑은 고딕', size=9, color=trend_color)
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 행 높이
    ws.row_dimensions[start_row].height = 25
    ws.row_dimensions[start_row + 1].height = 30
    ws.row_dimensions[start_row + 2].height = 30
    ws.row_dimensions[start_row + 3].height = 20
    
    return start_row + 4


def add_scorecard_table(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    data: list,
    title: str = "KPI Scorecard"
) -> int:
    """KPI 스코어카드 테이블
    
    Args:
        data: [{'name': str, 'actual': float, 'target': float, 'status': str}, ...]
    
    Returns: 종료 행 번호
    """
    sm = StyleManager()
    
    # 제목
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = title
    title_cell.font = sm.get_title_font()
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 4
    )
    ws.row_dimensions[start_row].height = 30
    
    # 헤더
    headers = ['KPI', 'Actual', 'Target', 'Status', 'Gauge']
    header_row = start_row + 1
    for i, header in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i)
        cell.value = header
        cell.font = sm.get_header_font()
        cell.fill = sm.get_header_fill()
        cell.alignment = sm.get_center_alignment()
        cell.border = sm.get_thin_border()
    ws.row_dimensions[header_row].height = 25
    
    # 데이터
    for row_idx, item in enumerate(data):
        row = header_row + 1 + row_idx
        
        # KPI 이름
        ws.cell(row=row, column=start_col).value = item.get('name', '')
        ws.cell(row=row, column=start_col).font = sm.get_data_font()
        ws.cell(row=row, column=start_col).alignment = Alignment(horizontal='left', vertical='center')
        
        # Actual
        actual = item.get('actual', 0)
        ws.cell(row=row, column=start_col + 1).value = actual
        ws.cell(row=row, column=start_col + 1).number_format = '#,##0.0'
        
        # Target
        target = item.get('target', 0)
        ws.cell(row=row, column=start_col + 2).value = target
        ws.cell(row=row, column=start_col + 2).number_format = '#,##0.0'
        
        # Status
        status = item.get('status', 'Unknown')
        status_cell = ws.cell(row=row, column=start_col + 3)
        status_cell.value = status
        status_cell.fill = sm.get_status_fill(status)
        status_cell.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
        
        # Gauge (텍스트)
        progress = actual / target if target > 0 else 0
        filled = int(min(progress, 1) * 10)
        gauge = '█' * filled + '░' * (10 - filled)
        gauge_cell = ws.cell(row=row, column=start_col + 4)
        gauge_cell.value = gauge
        gauge_cell.font = Font(name='Consolas', size=10)
        
        # 스타일
        for col in range(start_col, start_col + 5):
            ws.cell(row=row, column=col).border = sm.get_thin_border()
            ws.cell(row=row, column=col).alignment = sm.get_center_alignment()
        
        ws.row_dimensions[row].height = 22
    
    # 열 너비
    widths = [25, 12, 12, 15, 15]
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = width
    
    return header_row + len(data) + 1


def write_styled_dataframe(
    ws: Worksheet,
    df,  # pandas DataFrame
    start_row: int = 1,
    start_col: int = 1,
    title: str = None,
    number_format: str = '#,##0',
    add_heatmap: bool = False,
    heatmap_cols: list = None
) -> int:
    """스타일이 적용된 DataFrame 출력
    
    Returns: 종료 행 번호
    """
    sm = StyleManager()
    current_row = start_row
    
    # 제목
    if title:
        title_cell = ws.cell(row=current_row, column=start_col)
        title_cell.value = title
        title_cell.font = sm.get_title_font()
        ws.merge_cells(
            start_row=current_row, start_column=start_col,
            end_row=current_row, end_column=start_col + len(df.columns) - 1
        )
        ws.row_dimensions[current_row].height = 35
        current_row += 2
    
    # 헤더
    header_row = current_row
    for col_idx, col_name in enumerate(df.columns):
        cell = ws.cell(row=header_row, column=start_col + col_idx)
        cell.value = col_name
        cell.font = sm.get_header_font()
        cell.fill = sm.get_header_fill()
        cell.alignment = sm.get_center_alignment(wrap=True)
        cell.border = sm.get_thin_border()
    ws.row_dimensions[header_row].height = 30
    
    # 데이터
    for row_idx, row_data in enumerate(df.values):
        row = header_row + 1 + row_idx
        alt_fill = sm.get_data_fill(row_idx % 2 == 0)
        
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row, column=start_col + col_idx)
            cell.value = value
            cell.fill = alt_fill
            cell.border = sm.get_thin_border()
            
            # 숫자 포맷
            if isinstance(value, (int, float)):
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell.font = sm.get_data_font()
        
        ws.row_dimensions[row].height = 22
    
    # 열 너비 자동 조정
    for col_idx, col_name in enumerate(df.columns):
        max_length = len(str(col_name))
        for row_idx in range(len(df)):
            cell_value = df.iloc[row_idx, col_idx]
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 4, 30)
        ws.column_dimensions[get_column_letter(start_col + col_idx)].width = adjusted_width
    
    # 히트맵
    if add_heatmap and heatmap_cols:
        for col_name in heatmap_cols:
            if col_name in df.columns:
                col_idx = list(df.columns).index(col_name)
                add_heatmap_formatting(
                    ws,
                    header_row + 1,
                    header_row + len(df),
                    start_col + col_idx,
                    start_col + col_idx
                )
    
    return header_row + len(df) + 1


if __name__ == "__main__":
    # 테스트
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # KPI 카드 테스트
    add_kpi_card(ws, 1, 1, "Total Revenue", "₩149.2조", "YoY Growth", "↑", "+5.2%")
    add_kpi_card(ws, 1, 5, "Declarations", "21.4M", "vs Last Year", "↑", "+3.1%")
    
    # 리스크 매트릭스 테스트
    add_risk_matrix(ws, 7, 1, "Risk Assessment Matrix")
    
    # 스코어카드 테스트
    scorecard_data = [
        {'name': 'YoY Growth Rate', 'actual': 5.2, 'target': 5.0, 'status': 'Excellent'},
        {'name': 'Underval Rate', 'actual': 4.5, 'target': 3.0, 'status': 'Warning'},
        {'name': 'Compliance Rate', 'actual': 92.1, 'target': 95.0, 'status': 'Good'},
    ]
    add_scorecard_table(ws, 18, 1, scorecard_data, "KPI Performance")
    
    wb.save("/tmp/viz_test.xlsx")
    print("✅ Test file saved to /tmp/viz_test.xlsx")
