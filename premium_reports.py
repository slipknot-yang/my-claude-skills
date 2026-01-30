#!/usr/bin/env python3
"""
프리미엄 관세 분석 보고서 생성기

WCO PMM / KCS 관세연감 / UN Comtrade 수준의 보고서 생성

Features:
- Executive Summary (경영진 대시보드)
- WCO PMM KPI Scorecard
- Advanced Visualizations (Pareto, Heatmap, Risk Matrix)
- Professional Formatting
- Methodology & Glossary

Usage:
    python premium_reports.py
"""

import oracledb
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys

# 로컬 모듈 임포트
from kpi_calculator import KPICalculator, KPI_DEFINITIONS, KPICategory, format_currency, format_percent
from visualizations import (
    ColorPalette, StyleManager, 
    add_kpi_card, add_risk_matrix, add_scorecard_table,
    add_heatmap_formatting, add_databar_formatting,
    add_pareto_chart, add_combo_chart,
    write_styled_dataframe, get_trend_arrow
)

# === 설정 ===
DB_CONFIG = {
    "user": "CLRIUSR",
    "password": "ntancisclri1!",
    "dsn": "211.239.120.42:3535/NTANCIS"
}

BASE_PATH = os.path.dirname(os.path.abspath(__file__))


class PremiumReportGenerator:
    """프리미엄 관세 보고서 생성기"""
    
    def __init__(self, conn):
        self.conn = conn
        self.kpi_calc = KPICalculator(conn)
        self.sm = StyleManager()
        self.report_date = datetime.now().strftime('%Y-%m-%d')
    
    # === 공통 헬퍼 함수 ===
    
    def _create_cover_sheet(self, wb: Workbook, title: str, subtitle: str, metrics: dict) -> Worksheet:
        """표지 시트 생성 (UN Comtrade 스타일)"""
        ws = wb.create_sheet("Cover", 0)
        
        # 배경색
        for row in range(1, 35):
            for col in range(1, 16):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'
                )
        
        # 상단 배너 (파란색 그라데이션 효과)
        for col in range(1, 16):
            for row in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=ColorPalette.PRIMARY, 
                    end_color=ColorPalette.PRIMARY, 
                    fill_type='solid'
                )
        
        # 로고 텍스트
        ws.merge_cells('B2:N2')
        logo_cell = ws['B2']
        logo_cell.value = "KOREA CUSTOMS SERVICE"
        logo_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        logo_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 메인 타이틀
        ws.merge_cells('B6:N6')
        title_cell = ws['B6']
        title_cell.value = title
        title_cell.font = Font(name='맑은 고딕', size=32, bold=True, color=ColorPalette.PRIMARY)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[6].height = 60
        
        # 서브타이틀
        ws.merge_cells('B8:N8')
        sub_cell = ws['B8']
        sub_cell.value = subtitle
        sub_cell.font = Font(name='맑은 고딕', size=14, color=ColorPalette.SECONDARY)
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 날짜
        ws.merge_cells('B10:N10')
        date_cell = ws['B10']
        date_cell.value = f"Report Date: {self.report_date}"
        date_cell.font = Font(name='맑은 고딕', size=11, color=ColorPalette.DARK_GRAY)
        date_cell.alignment = Alignment(horizontal='center')
        
        # KPI 카드들 (4열)
        row = 14
        col_positions = [2, 5, 8, 11]
        
        for i, (label, value) in enumerate(metrics.items()):
            if i >= 4:
                break
            col = col_positions[i]
            
            # 카드 배경
            for r in range(row, row + 4):
                for c in range(col, col + 3):
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'
                    )
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
            
            # 라벨
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
            label_cell = ws.cell(row=row, column=col)
            label_cell.value = label
            label_cell.font = Font(name='맑은 고딕', size=10, color=ColorPalette.DARK_GRAY)
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 값
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+2)
            value_cell = ws.cell(row=row+1, column=col)
            value_cell.value = value
            value_cell.font = Font(name='맑은 고딕', size=22, bold=True, color=ColorPalette.PRIMARY)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 하단 정보
        ws.merge_cells('B30:N30')
        footer = ws['B30']
        footer.value = "Data Source: CLRI_TANSAD_ITM_D, CLRI_TANSAD_UT_PRC_M | Methodology: WCO PMM Framework"
        footer.font = Font(name='맑은 고딕', size=9, color=ColorPalette.DARK_GRAY)
        footer.alignment = Alignment(horizontal='center')
        
        # 열 너비
        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 10
        
        return ws
    
    def _create_executive_summary(self, wb: Workbook) -> Worksheet:
        """Executive Summary 시트 (경영진 대시보드)"""
        ws = wb.create_sheet("Executive Summary")
        
        # 요약 데이터
        summary = self.kpi_calc.calc_executive_summary()
        scorecard = self.kpi_calc.calc_kpi_scorecard()
        
        # 제목
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "Executive Dashboard"
        title_cell.font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws.row_dimensions[1].height = 40
        
        # 부제목
        ws.merge_cells('A2:L2')
        ws['A2'].value = f"Analysis Period: {summary.get('period', 'N/A')} | Generated: {self.report_date}"
        ws['A2'].font = Font(name='맑은 고딕', size=10, color=ColorPalette.DARK_GRAY)
        
        # KPI 카드들 (1행에 4개)
        row = 4
        cards = [
            ("Total Declarations", f"{summary['total_declarations']:,.0f}", "건수"),
            ("Total Tax Revenue", format_currency(summary['total_tax_krw'], 'KRW'), "관세수입"),
            ("Total Import Value", format_currency(summary['total_value_usd'], 'USD'), "수입액"),
            ("YoY Growth", f"{summary['yoy_growth_pct']:+.1f}%", "성장률"),
        ]
        
        for i, (label, value, sub) in enumerate(cards):
            col = 1 + i * 3
            end_row = add_kpi_card(ws, row, col, label, value, sub)
        
        # KPI Scorecard 테이블
        scorecard_data = []
        for _, row_data in scorecard.iterrows():
            scorecard_data.append({
                'name': row_data['name_kr'],
                'actual': row_data['actual'],
                'target': row_data['target'] if pd.notna(row_data['target']) else 0,
                'status': row_data['status']
            })
        
        add_scorecard_table(ws, 10, 1, scorecard_data, "WCO PMM KPI Scorecard")
        
        # 리스크 매트릭스
        add_risk_matrix(ws, 10, 8, "Risk Assessment Framework")
        
        # Key Findings 섹션
        findings_row = 22
        ws.merge_cells(f'A{findings_row}:L{findings_row}')
        ws[f'A{findings_row}'].value = "Key Findings & Recommendations"
        ws[f'A{findings_row}'].font = Font(name='맑은 고딕', size=14, bold=True, color=ColorPalette.PRIMARY)
        
        findings = [
            f"• Total customs revenue of {format_currency(summary['total_tax_krw'], 'KRW')} collected from {summary['total_declarations']:,.0f} declarations",
            f"• Year-over-year growth rate of {summary['yoy_growth_pct']:+.1f}% indicates {'positive' if summary['yoy_growth_pct'] > 0 else 'negative'} trend",
            f"• Commodity HHI of {summary['hhi_commodity']:.0f} suggests {'moderate' if summary['hhi_commodity'] < 1800 else 'high'} concentration",
            f"• Undervaluation rate of {summary['underval_rate']:.1f}% detected in recent declarations",
            f"• Recommendation: Focus audit resources on high-risk HS-Country combinations"
        ]
        
        for i, finding in enumerate(findings):
            ws[f'A{findings_row + 1 + i}'].value = finding
            ws[f'A{findings_row + 1 + i}'].font = Font(name='맑은 고딕', size=10)
        
        return ws
    
    def _create_methodology_sheet(self, wb: Workbook) -> Worksheet:
        """Methodology 시트"""
        ws = wb.create_sheet("Methodology")
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'].value = "Data Sources & Methodology"
        ws['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws.row_dimensions[1].height = 40
        
        # 데이터 소스
        ws['A3'].value = "1. Data Sources"
        ws['A3'].font = Font(name='맑은 고딕', size=14, bold=True)
        
        sources = [
            ("CLRI_TANSAD_ITM_D", "Import declaration item data", "~21M records"),
            ("CLRI_TANSAD_UT_PRC_M", "Unit price assessment data", "~14M records"),
        ]
        
        for i, (table, desc, size) in enumerate(sources):
            ws[f'A{5+i}'].value = f"  • {table}"
            ws[f'A{5+i}'].font = Font(name='Consolas', size=10)
            ws[f'D{5+i}'].value = desc
            ws[f'G{5+i}'].value = size
        
        # KPI 계산 방법론
        ws['A9'].value = "2. KPI Calculation Methodology"
        ws['A9'].font = Font(name='맑은 고딕', size=14, bold=True)
        
        kpi_methods = [
            ("YoY Growth Rate", "(Current Year Tax - Previous Year Tax) / Previous Year Tax × 100"),
            ("HHI Index", "Sum of squared market shares × 10,000"),
            ("Undervaluation Rate", "Count(Assessed > Declared × 1.3) / Total Count × 100"),
            ("Volatility (CV)", "Standard Deviation / Mean × 100"),
        ]
        
        for i, (kpi, formula) in enumerate(kpi_methods):
            ws[f'A{11+i}'].value = f"  • {kpi}:"
            ws[f'A{11+i}'].font = Font(name='맑은 고딕', size=10, bold=True)
            ws[f'D{11+i}'].value = formula
            ws[f'D{11+i}'].font = Font(name='Consolas', size=9)
        
        # 참조 프레임워크
        ws['A17'].value = "3. Reference Frameworks"
        ws['A17'].font = Font(name='맑은 고딕', size=14, bold=True)
        
        refs = [
            "• WCO Performance Measurement Model (PMM) - 4 Dimensions",
            "• WCO Customs Risk Management Compendium (2022)",
            "• UN Comtrade Database Analytical Standards",
            "• Korea Customs Service Annual Statistics Methodology",
        ]
        
        for i, ref in enumerate(refs):
            ws[f'A{19+i}'].value = ref
            ws[f'A{19+i}'].font = Font(name='맑은 고딕', size=10)
        
        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['G'].width = 15
        
        return ws
    
    def _create_glossary_sheet(self, wb: Workbook) -> Worksheet:
        """Glossary 시트 (용어 정의 한/영)"""
        ws = wb.create_sheet("Glossary")
        
        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'].value = "Glossary / 용어 정의"
        ws['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws.row_dimensions[1].height = 40
        
        # 헤더
        headers = ['Term (EN)', 'Term (KR)', 'Abbreviation', 'Definition']
        for i, h in enumerate(headers):
            cell = ws.cell(row=3, column=i+1)
            cell.value = h
            cell.font = self.sm.get_header_font()
            cell.fill = self.sm.get_header_fill()
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 용어 목록
        glossary = [
            ("Customs Duty", "관세", "CD", "Tax imposed on imported goods"),
            ("HS Code", "품목분류코드", "HS", "Harmonized System commodity classification code"),
            ("Declared Value", "신고가격", "DV", "Value declared by importer"),
            ("Assessed Value", "심사가격", "AV", "Value determined by customs"),
            ("Undervaluation", "과소신고", "-", "Declaring lower value than actual"),
            ("HHI", "허핀달-허쉬만지수", "HHI", "Market concentration index (0-10,000)"),
            ("YoY", "전년대비", "YoY", "Year-over-Year comparison"),
            ("MoM", "전월대비", "MoM", "Month-over-Month comparison"),
            ("WCO", "세계관세기구", "WCO", "World Customs Organization"),
            ("PMM", "성과측정모델", "PMM", "Performance Measurement Model"),
            ("KPI", "핵심성과지표", "KPI", "Key Performance Indicator"),
            ("Risk Matrix", "리스크 매트릭스", "-", "Frequency × Severity assessment grid"),
            ("Pareto Analysis", "파레토 분석", "-", "80/20 rule analysis"),
            ("Coefficient of Variation", "변동계수", "CV", "Standard deviation / Mean × 100"),
        ]
        
        for i, (en, kr, abbr, defn) in enumerate(glossary):
            row = 4 + i
            ws.cell(row=row, column=1).value = en
            ws.cell(row=row, column=2).value = kr
            ws.cell(row=row, column=3).value = abbr
            ws.cell(row=row, column=4).value = defn
            
            # 스타일
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name='맑은 고딕', size=10)
                cell.border = self.sm.get_thin_border()
                if col == 3:
                    cell.alignment = Alignment(horizontal='center')
        
        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 45
        
        return ws
    
    # === 관세 수입 현황 보고서 ===
    
    def create_revenue_report(self, output_path: str):
        """프리미엄 관세 수입 현황 보고서 생성"""
        print("📊 프리미엄 관세 수입 현황 보고서 생성 중...")
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # 데이터 조회
        print("  → 데이터 조회...")
        df_yearly = self.kpi_calc.calc_revenue_by_period('yearly')
        df_monthly = self.kpi_calc.calc_revenue_by_period('monthly')
        df_yoy = self.kpi_calc.calc_yoy_growth()
        df_pareto = self.kpi_calc.calc_pareto_analysis('hs2', 'tax')
        hhi_hs = self.kpi_calc.calc_hhi_by_dimension('hs2')
        hhi_country = self.kpi_calc.calc_hhi_by_dimension('country')
        summary = self.kpi_calc.calc_executive_summary()
        
        # 국가별 데이터
        df_country = pd.read_sql("""
            SELECT 
                ORIG_CNTY_CD as country,
                COUNT(*) as declarations,
                SUM(ITM_TAX_AMT) as tax_amount,
                SUM(ITM_INVC_USD_AMT) as value_usd,
                ROUND(SUM(ITM_TAX_AMT) * 100.0 / SUM(SUM(ITM_TAX_AMT)) OVER(), 2) as share_pct
            FROM CLRI_TANSAD_ITM_D
            WHERE DEL_YN = 'N' AND TANSAD_YY >= '23' AND ORIG_CNTY_CD IS NOT NULL
            GROUP BY ORIG_CNTY_CD
            ORDER BY tax_amount DESC
            FETCH FIRST 20 ROWS ONLY
        """, self.conn)
        df_country.columns = ['Country', 'Declarations', 'Tax Amount', 'Value (USD)', 'Share %']
        
        # 표지 생성
        print("  → 표지 생성...")
        metrics = {
            'Total Declarations': f"{summary['total_declarations']:,.0f}",
            'Total Tax Revenue': format_currency(summary['total_tax_krw'], 'KRW'),
            'Total Import Value': format_currency(summary['total_value_usd'], 'USD'),
            'Analysis Period': summary['period']
        }
        self._create_cover_sheet(wb, "Customs Revenue Analysis", "관세 수입 현황 분석 보고서", metrics)
        
        # Executive Summary
        print("  → Executive Summary 생성...")
        self._create_executive_summary(wb)
        
        # 연도별 추이 시트
        print("  → 연도별 추이 시트 생성...")
        ws_yearly = wb.create_sheet("Yearly Trend")
        
        # 제목
        ws_yearly.merge_cells('A1:H1')
        ws_yearly['A1'].value = "Yearly Revenue Trend Analysis"
        ws_yearly['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        ws_yearly.row_dimensions[1].height = 35
        
        # 데이터 테이블
        end_row = write_styled_dataframe(ws_yearly, df_yoy, start_row=3, title="Annual Revenue with YoY Growth")
        
        # 콤보 차트 (세액 + 성장률)
        if len(df_yoy) > 1:
            add_combo_chart(
                ws_yearly,
                bar_col=3,  # total_tax
                line_col=6,  # yoy_growth_pct
                cat_col=1,   # period
                start_row=3,
                end_row=3 + len(df_yoy),
                position="I3",
                title="Revenue & Growth Trend",
                bar_title="Tax Revenue",
                line_title="YoY Growth %"
            )
        
        # 품목별 파레토 분석 시트
        print("  → 파레토 분석 시트 생성...")
        ws_pareto = wb.create_sheet("Pareto Analysis")
        
        ws_pareto.merge_cells('A1:H1')
        ws_pareto['A1'].value = "Pareto Analysis (80/20 Rule)"
        ws_pareto['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        # 파레토 테이블
        pareto_display = df_pareto[['category', 'value', 'share_pct', 'cumulative_pct', 'pareto_zone']].head(20)
        pareto_display.columns = ['HS Chapter', 'Tax Amount', 'Share %', 'Cumulative %', 'Zone']
        end_row = write_styled_dataframe(ws_pareto, pareto_display, start_row=3, title="Top 20 HS Chapters by Tax Revenue")
        
        # 히트맵 서식 (Share % 컬럼)
        add_heatmap_formatting(ws_pareto, 6, 5 + len(pareto_display), 3, 3)
        
        # Zone A 카운트 요약
        zone_a = len(df_pareto[df_pareto['pareto_zone'] == 'A (Top 80%)'])
        ws_pareto[f'A{end_row + 2}'].value = f"※ Zone A (Top 80%): {zone_a} HS chapters account for 80% of revenue"
        ws_pareto[f'A{end_row + 2}'].font = Font(name='맑은 고딕', size=11, bold=True, color=ColorPalette.PRIMARY)
        
        # HHI 요약
        ws_pareto[f'A{end_row + 4}'].value = f"HHI Concentration Index: {hhi_hs['hhi']:.0f} ({hhi_hs['concentration_level']})"
        ws_pareto[f'A{end_row + 5}'].value = f"Top 5 Share: {hhi_hs['top_5_share']:.1f}%"
        
        # 국가별 현황 시트
        print("  → 국가별 현황 시트 생성...")
        ws_country = wb.create_sheet("Country Analysis")
        
        ws_country.merge_cells('A1:H1')
        ws_country['A1'].value = "Country of Origin Analysis"
        ws_country['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        end_row = write_styled_dataframe(ws_country, df_country, start_row=3, title="Top 20 Countries by Tax Revenue")
        
        # 데이터바 (Share % 컬럼)
        add_databar_formatting(ws_country, 6, 5 + len(df_country), 5)
        
        # HHI 국가
        ws_country[f'A{end_row + 2}'].value = f"HHI Concentration Index: {hhi_country['hhi']:.0f} ({hhi_country['concentration_level']})"
        
        # 월별 추이 시트
        print("  → 월별 추이 시트 생성...")
        ws_monthly = wb.create_sheet("Monthly Trend")
        
        ws_monthly.merge_cells('A1:H1')
        ws_monthly['A1'].value = "Monthly Revenue Trend (Last 36 Months)"
        ws_monthly['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        end_row = write_styled_dataframe(ws_monthly, df_monthly.head(36), start_row=3, title="Monthly Revenue Data")
        
        # 라인 차트
        chart = LineChart()
        chart.style = 10
        chart.title = "Monthly Tax Revenue Trend"
        chart.y_axis.title = "Tax Amount"
        
        data = Reference(ws_monthly, min_col=3, min_row=5, max_row=5 + min(36, len(df_monthly)))
        cats = Reference(ws_monthly, min_col=1, min_row=6, max_row=5 + min(36, len(df_monthly)))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        if chart.series:
            chart.series[0].graphicalProperties.line.solidFill = ColorPalette.PRIMARY
            chart.series[0].graphicalProperties.line.width = 25000
            chart.series[0].smooth = True
        
        chart.width = 18
        chart.height = 10
        ws_monthly.add_chart(chart, "F3")
        
        # Methodology & Glossary
        print("  → 부록 시트 생성...")
        self._create_methodology_sheet(wb)
        self._create_glossary_sheet(wb)
        
        # 저장
        wb.save(output_path)
        print(f"✅ 저장 완료: {output_path}")
    
    # === 이상 탐지 보고서 ===
    
    def create_anomaly_report(self, output_path: str):
        """프리미엄 이상 탐지 보고서 생성"""
        print("\n🚨 프리미엄 이상 탐지 보고서 생성 중...")
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # 데이터 조회
        print("  → 데이터 조회...")
        df_underval = self.kpi_calc.calc_undervaluation_stats()
        df_risk = self.kpi_calc.calc_risk_score_by_hs_country()
        df_misclass = self.kpi_calc.calc_hs_misclassification_rate()
        
        # 고위험 업체
        df_importers = pd.read_sql("""
            SELECT 
                IMPPN_TIN as tin,
                MAX(IMPPN_NM) as importer_name,
                COUNT(*) as total,
                SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 
                         AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) as underval,
                ROUND(SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 
                               AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) * 100.0 / COUNT(*), 1) as rate_pct,
                SUM(ASSD_INVC_USD_AMT) as total_value
            FROM CLRI_TANSAD_UT_PRC_M
            WHERE DEL_YN = 'N' AND TANSAD_YY >= '23'
            GROUP BY IMPPN_TIN
            HAVING COUNT(*) >= 50
               AND SUM(CASE WHEN ASSD_UT_USD_VAL > DCLD_UT_USD_VAL * 1.3 
                            AND DCLD_UT_USD_VAL > 0 THEN 1 ELSE 0 END) >= 10
            ORDER BY underval DESC
            FETCH FIRST 30 ROWS ONLY
        """, self.conn)
        df_importers.columns = ['TIN', 'Importer Name', 'Total', 'Underval', 'Rate %', 'Total Value']
        
        # 요약 통계
        total_underval = df_underval['underval_count'].sum() if len(df_underval) > 0 else 0
        total_loss = df_underval['estimated_loss_usd'].sum() if len(df_underval) > 0 else 0
        high_risk_combos = len(df_risk[df_risk['risk_score'] >= 50]) if len(df_risk) > 0 else 0
        
        # 표지
        print("  → 표지 생성...")
        metrics = {
            'Underval Cases': f"{total_underval:,.0f}",
            'Est. Revenue Loss': format_currency(total_loss, 'USD'),
            'High-Risk Combos': f"{high_risk_combos}",
            'At-Risk Importers': f"{len(df_importers)}"
        }
        self._create_cover_sheet(wb, "Anomaly Detection Report", "관세 이상 탐지 분석 보고서", metrics)
        
        # Executive Summary with Risk Matrix
        print("  → Executive Summary 생성...")
        ws_exec = wb.create_sheet("Risk Overview")
        
        ws_exec.merge_cells('A1:L1')
        ws_exec['A1'].value = "Risk Assessment Dashboard"
        ws_exec['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=ColorPalette.PRIMARY)
        ws_exec.row_dimensions[1].height = 40
        
        # KPI 카드들
        cards = [
            ("Undervaluation Cases", f"{total_underval:,.0f}", "30%+ threshold"),
            ("Estimated Loss", format_currency(total_loss, 'USD'), "potential duty"),
            ("High-Risk Combos", f"{high_risk_combos}", "score >= 50"),
            ("At-Risk Importers", f"{len(df_importers)}", "repeat offenders"),
        ]
        
        for i, (label, value, sub) in enumerate(cards):
            col = 1 + i * 3
            add_kpi_card(ws_exec, 3, col, label, value, sub)
        
        # 리스크 매트릭스
        add_risk_matrix(ws_exec, 9, 1, "Risk Assessment Matrix")
        
        # 요약 통계 테이블
        underval_summary = df_underval[['period', 'total_count', 'underval_count', 'underval_rate']].head(5)
        underval_summary.columns = ['Year', 'Total', 'Underval', 'Rate %']
        write_styled_dataframe(ws_exec, underval_summary, start_row=9, start_col=8, title="Undervaluation Trend by Year")
        
        # 과소신고 분석 시트
        print("  → 과소신고 분석 시트 생성...")
        ws_underval = wb.create_sheet("Undervaluation Analysis")
        
        ws_underval.merge_cells('A1:H1')
        ws_underval['A1'].value = "Undervaluation Detection Analysis"
        ws_underval['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        underval_display = df_underval.copy()
        underval_display.columns = ['Year', 'Total Count', 'Underval Count', 'Underval Rate %', 'Est. Loss (USD)']
        end_row = write_styled_dataframe(ws_underval, underval_display, start_row=3, title="Undervaluation Statistics by Year")
        
        # 히트맵 (Rate 컬럼)
        add_heatmap_formatting(ws_underval, 6, 5 + len(underval_display), 4, 4, reverse=True)
        
        # 바 차트
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Undervaluation Rate by Year"
        
        data = Reference(ws_underval, min_col=4, min_row=5, max_row=5 + len(underval_display))
        cats = Reference(ws_underval, min_col=1, min_row=6, max_row=5 + len(underval_display))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = ColorPalette.DANGER
        
        chart.width = 12
        chart.height = 8
        ws_underval.add_chart(chart, "G3")
        
        # HS-Country 리스크 시트
        print("  → HS-Country 리스크 시트 생성...")
        ws_risk = wb.create_sheet("HS-Country Risk")
        
        ws_risk.merge_cells('A1:I1')
        ws_risk['A1'].value = "HS Code × Country Risk Analysis"
        ws_risk['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        risk_display = df_risk[['hs4', 'country', 'total_count', 'underval_count', 'underval_rate', 'risk_score']].head(30)
        risk_display.columns = ['HS Code', 'Country', 'Total', 'Underval', 'Rate %', 'Risk Score']
        end_row = write_styled_dataframe(ws_risk, risk_display, start_row=3, title="Top 30 High-Risk HS-Country Combinations")
        
        # 리스크 점수 히트맵
        add_heatmap_formatting(ws_risk, 6, 5 + len(risk_display), 6, 6, reverse=True)
        
        # 고위험 업체 시트
        print("  → 고위험 업체 시트 생성...")
        ws_importers = wb.create_sheet("At-Risk Importers")
        
        ws_importers.merge_cells('A1:G1')
        ws_importers['A1'].value = "High-Risk Importer Analysis"
        ws_importers['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        end_row = write_styled_dataframe(ws_importers, df_importers, start_row=3, title="Top 30 At-Risk Importers (by Undervaluation Count)")
        
        # Rate % 히트맵
        add_heatmap_formatting(ws_importers, 6, 5 + len(df_importers), 5, 5, reverse=True)
        
        # HS 분류 오류 시트
        print("  → HS 분류 오류 시트 생성...")
        ws_misclass = wb.create_sheet("HS Misclassification")
        
        ws_misclass.merge_cells('A1:E1')
        ws_misclass['A1'].value = "HS Code Misclassification Analysis"
        ws_misclass['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color=ColorPalette.PRIMARY)
        
        misclass_display = df_misclass.copy()
        misclass_display.columns = ['Year', 'Total Count', 'Misclass Count', 'Misclass Rate %']
        end_row = write_styled_dataframe(ws_misclass, misclass_display, start_row=3, title="HS Misclassification by Year")
        
        # 부록
        print("  → 부록 시트 생성...")
        self._create_methodology_sheet(wb)
        self._create_glossary_sheet(wb)
        
        # 저장
        wb.save(output_path)
        print(f"✅ 저장 완료: {output_path}")


def main():
    """메인 실행"""
    print("=" * 60)
    print("🚀 프리미엄 관세 분석 보고서 생성")
    print(f"⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # DB 연결
    print("\n🔗 DB 연결 중...")
    try:
        conn = oracledb.connect(**DB_CONFIG)
        print("✅ DB 연결 성공")
    except Exception as e:
        print(f"❌ DB 연결 실패: {e}")
        sys.exit(1)
    
    try:
        generator = PremiumReportGenerator(conn)
        
        # 관세 수입 현황 보고서
        revenue_path = os.path.join(BASE_PATH, "프리미엄_관세수입현황_보고서.xlsx")
        generator.create_revenue_report(revenue_path)
        
        # 이상 탐지 보고서
        anomaly_path = os.path.join(BASE_PATH, "프리미엄_이상탐지_보고서.xlsx")
        generator.create_anomaly_report(anomaly_path)
        
        print("\n" + "=" * 60)
        print("✅ 모든 프리미엄 보고서 생성 완료!")
        print(f"📁 관세 수입 현황: {revenue_path}")
        print(f"📁 이상 탐지: {anomaly_path}")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()
        print("🔌 DB 연결 종료")


if __name__ == "__main__":
    main()
